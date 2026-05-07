"""Financial dashboard routes."""

import csv
import html
import hashlib
import io
import json
import logging
import math
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, timezone
from threading import Lock
from typing import Any, cast
from uuid import uuid4

import requests as http_requests
from flask import Flask, jsonify, render_template, request
from flask_limiter import Limiter

from .cache import get_cache
from .finance_cashflow_features import (
    SmartDedupeCache,
    apply_bulk_cashflow_updates,
    bulk_delete_cashflow_entries,
    evaluate_data_quality_alerts,
    validate_bulk_operation_ids,
)
from .finance_cashflow_helpers import (
    build_reconcile_suggestions,
    cashflow_dedupe_hash,
    find_potential_cashflow_duplicate,
    normalize_cashflow_text,
    parse_cashflow_import_candidates,
    tokenize_cashflow_text,
)
from .finance_blueprints.analytics import register_analytics_routes
from .finance_blueprints.anomalies import register_anomaly_routes
from .finance_blueprints.accounts import register_account_routes
from .finance_blueprints.assets import register_assets_routes
from .finance_blueprints.admin import register_admin_routes
from .finance_blueprints.cashflow import register_cashflow_routes
from .finance_blueprints.credit_cards import register_credit_card_routes
from .finance_blueprints.debts import register_debt_routes
from .finance_blueprints.dividends import register_dividends_routes
from .finance_blueprints.goals import register_goals_routes
from .finance_blueprints.insights import register_insights_routes
from .finance_blueprints.investment_planning import register_investment_planning_routes
from .finance_blueprints.maintenance import register_maintenance_routes
from .finance_blueprints.monthly_comparison import register_monthly_comparison_routes
from .finance_blueprints.performance import register_performance_routes
from .finance_blueprints.security import register_security_routes
from .finance_blueprints.watchlist import register_watchlist_routes
from .repository import Repository
from .security import require_finance_key, sanitize_text

_RETRY_EXCEPTIONS = (
    http_requests.exceptions.ConnectionError,
    http_requests.exceptions.Timeout,
    http_requests.exceptions.ChunkedEncodingError,
)

def _build_simple_pdf(lines: list[str], theme: str = "light") -> bytes:
    """Build a tiny text-only PDF without external dependencies."""
    safe_lines = [
        str(line or "").replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        for line in lines
    ]
    dark_mode = str(theme or "light").strip().lower() == "dark"
    y = 800
    content_lines = []
    if dark_mode:
        content_lines.extend([
            "0 0 0 rg",
            "0 0 595 842 re f",
        ])
    content_lines.extend([
        "BT",
        "/F1 11 Tf",
        "1 1 1 rg" if dark_mode else "0 0 0 rg",
        "1 0 0 1 40 800 Tm",
    ])
    for line in safe_lines:
        content_lines.append(f"1 0 0 1 40 {y} Tm ({line}) Tj")
        y -= 14
        if y < 40:
            break
    content_lines.append("ET")
    stream = "\n".join(content_lines).encode("latin-1", errors="replace")

    objs: list[bytes] = []
    objs.append(b"<< /Type /Catalog /Pages 2 0 R >>")
    objs.append(b"<< /Type /Pages /Count 1 /Kids [3 0 R] >>")
    objs.append(
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
        b"/Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>"
    )
    objs.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    objs.append(b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream")

    out = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for i, obj in enumerate(objs, start=1):
        offsets.append(len(out))
        out.extend(f"{i} 0 obj\n".encode("ascii"))
        out.extend(obj)
        out.extend(b"\nendobj\n")

    xref_pos = len(out)
    out.extend(f"xref\n0 {len(objs) + 1}\n".encode("ascii"))
    out.extend(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        out.extend(f"{off:010d} 00000 n \n".encode("ascii"))
    out.extend(
        (
            f"trailer\n<< /Size {len(objs) + 1} /Root 1 0 R >>\n"
            f"startxref\n{xref_pos}\n%%EOF"
        ).encode("ascii")
    )
    return bytes(out)

def _http_get_with_retry(
    url: str,
    *,
    params: dict | None = None,
    headers: dict | None = None,
    timeout: int = 10,
    max_retries: int = 2,
) -> http_requests.Response:
    """GET with exponential-backoff retry on transient network errors.

    Only retries on connection/timeout failures — never on 4xx/5xx responses.
    """
    delay = 0.5
    for attempt in range(max_retries + 1):
        try:
            return http_requests.get(
                url, params=params, headers=headers, timeout=timeout,
            )
        except _RETRY_EXCEPTIONS as exc:
            if attempt >= max_retries:
                raise
            logging.getLogger(__name__).debug(
                "Retry %d/%d for %s after error: %s",
                attempt + 1, max_retries, url, exc,
            )
            time.sleep(delay)
            delay *= 2
    raise RuntimeError("unreachable")

def register_finance_routes(app: Flask, limiter: Limiter) -> None:
    logger = logging.getLogger(__name__)
    repo = Repository(
        app.config["DATABASE_TARGET"],
        enable_runtime_schema_evolution=app.config.get("ALLOW_RUNTIME_SCHEMA_EVOLUTION", True),
    )
    cache = get_cache(app.config)
    setattr(app, "_finance_cache", cache)  # exposed for tests

    # Initialize smart dedup cache (1 hour TTL, Redis with fallback to memory)
    dedup_ttl = int(app.config.get("CASHFLOW_DEDUP_TTL_SECONDS", 3600))
    _dedupe_cache = SmartDedupeCache(
        backend_cache=cache,
        ttl_seconds=dedup_ttl,
        prefix="cashflow:dedup:",
    )

    FINANCE_CACHE_TTLS = {
        "summary": 60,
        "cashflow_summary": 30,
        "cashflow_analytics": 30,
        "market": 300,
        "portfolio_history": 120,
        "asset_history": 120,
        "invested_history": 120,
        "performance": 90,
        "audit": 30,
        "allocation_targets": 120,
        "dividend_summary": 120,
        "passive_income_goal": 300,
        "dividend_ceiling": 300,
        "independence": 180,
        "benchmark": 600,
    }
    CLEANUP_CONFIRM_TOKEN = "CONFIRM_CLEANUP_DUPLICATES"
    CLEANUP_COOLDOWN_SECONDS = 600
    CLEANUP_IDEMPOTENCY_TTL = 3600

    def _invalidate_cache_prefixes(*prefixes: str) -> None:
        for prefix in prefixes:
            if not prefix:
                continue
            if hasattr(cache, "delete_prefix"):
                cache.delete_prefix(prefix)
            else:
                cache.delete(prefix)

    def _invalidate_financial_state_cache(
        *,
        include_market: bool = False,
        include_dividends: bool = False,
    ) -> None:
        prefixes = [
            "finance:summary",
            "finance:portfolio-history:",
            "finance:asset-history:",
            "finance:invested-history:",
            "finance:performance:",
            "finance:dividend-ceiling:",
            "finance:independence:",
        ]
        if include_dividends:
            prefixes.append("finance:dividend-summary")
        if include_market:
            prefixes.append("finance:market")
        _invalidate_cache_prefixes(*prefixes)

    def _invalidate_cashflow_cache() -> None:
        _invalidate_cache_prefixes(
            "finance:cashflow-summary:",
            "finance:cashflow-analytics:",
        )

    CASHFLOW_RECONCILE_REVIEW_TTL_SECONDS = 15 * 60
    CASHFLOW_IMPORT_JOB_TTL_SECONDS = 30 * 60
    CASHFLOW_IMPORT_ASYNC_ROW_THRESHOLD = 350

    _reconcile_reviews: dict[str, dict] = {}
    _reconcile_reviews_lock = Lock()
    _cashflow_import_jobs: dict[str, dict] = {}
    _cashflow_import_jobs_lock = Lock()
    _cashflow_import_executor = ThreadPoolExecutor(max_workers=2)
    _ocr_cache: dict[str, dict] = {}
    _ocr_cache_lock = Lock()
    _ocr_history: list[dict] = []   # last 20 OCR results (in-memory)
    _ocr_history_lock = Lock()
    OCR_CACHE_TTL_SECONDS = 60 * 60  # 1 hour (was 10 min)

    def _cleanup_cashflow_review_cache() -> None:
        now_ts = time.time()
        with _reconcile_reviews_lock:
            expired = [
                k for k, v in _reconcile_reviews.items()
                if now_ts - float(v.get("created_ts") or 0.0) > CASHFLOW_RECONCILE_REVIEW_TTL_SECONDS
            ]
            for key in expired:
                _reconcile_reviews.pop(key, None)

    def _cleanup_cashflow_import_jobs() -> None:
        now_ts = time.time()
        with _cashflow_import_jobs_lock:
            expired = [
                k for k, v in _cashflow_import_jobs.items()
                if now_ts - float(v.get("created_ts") or 0.0) > CASHFLOW_IMPORT_JOB_TTL_SECONDS
            ]
            for key in expired:
                _cashflow_import_jobs.pop(key, None)

    def _cleanup_ocr_cache() -> None:
        now_ts = time.time()
        with _ocr_cache_lock:
            expired = [
                key for key, row in _ocr_cache.items()
                if now_ts - float(row.get("created_ts") or 0.0) > OCR_CACHE_TTL_SECONDS
            ]
            for key in expired:
                _ocr_cache.pop(key, None)

    def _build_reconcile_suggestions(
        month: str | None,
        min_score: float,
    ) -> list[dict]:
        rows = repo.list_fin_cashflow_entries(month=month or None, limit=5000)
        return build_reconcile_suggestions(rows, min_score)

    def _execute_cashflow_import(
        *,
        month: str,
        filename: str,
        raw_bytes: bytes,
        force: bool,
        merge_strategy: str,
    ) -> dict:
        candidates, parse_errors = parse_cashflow_import_candidates(filename, raw_bytes)
        existing_entries = repo.list_fin_cashflow_entries(month=month, limit=5000)

        imported = 0
        merged = 0
        inserted_ids: list[int] = []
        potential_duplicates: list[dict] = []

        for candidate in candidates:
            entry_date = str(candidate.get("entry_date") or "")[:10]
            if not re.match(r"^\d{4}-\d{2}-\d{2}$", entry_date):
                parse_errors.append({"error": "entry_date inválida", "entry_date": entry_date})
                continue

            if str(entry_date)[:7] != month:
                parse_errors.append({
                    "error": "entry_date fora do month informado",
                    "entry_date": entry_date,
                    "month": month,
                })
                continue

            duplicate = None
            if not force:
                duplicate = find_potential_cashflow_duplicate(
                    existing_entries=existing_entries,
                    entry_type=str(candidate.get("entry_type") or ""),
                    amount=float(candidate.get("amount") or 0),
                    entry_date=entry_date,
                    description=str(candidate.get("description") or ""),
                )

            if duplicate and float(duplicate.get("score") or 0) >= 60.0:
                potential_duplicates.append({
                    "incoming": {
                        "entry_type": candidate.get("entry_type"),
                        "amount": candidate.get("amount"),
                        "entry_date": entry_date,
                        "description": candidate.get("description"),
                    },
                    "existing": duplicate,
                })

                if merge_strategy == "append_notes":
                    existing_id = int(duplicate.get("id") or 0)
                    existing_row = repo.get_fin_cashflow_entry(existing_id) or {}
                    current_notes = str(existing_row.get("notes") or "").strip()
                    append_text = (
                        "Import dedupe: "
                        f"{entry_date} {str(candidate.get('description') or '').strip()}"
                    )
                    new_notes = (
                        f"{current_notes} | {append_text}" if current_notes else append_text
                    )[:500]
                    if repo.update_fin_cashflow_entry(existing_id, {"notes": new_notes}):
                        merged += 1
                continue

            entry_id = repo.add_fin_cashflow_entry(candidate)
            inserted_ids.append(int(entry_id))
            imported += 1
            existing_entries.append({**candidate, "id": entry_id})

        return {
            "ok": True,
            "month": month,
            "filename": filename,
            "imported": imported,
            "merged": merged,
            "inserted_ids": inserted_ids,
            "potential_duplicates": potential_duplicates,
            "errors": parse_errors,
        }

    def _enqueue_cashflow_import_job(
        *,
        month: str,
        filename: str,
        raw_bytes: bytes,
        force: bool,
        merge_strategy: str,
    ) -> str:
        _cleanup_cashflow_import_jobs()
        job_id = uuid4().hex
        now_iso = datetime.now(timezone.utc).isoformat()
        with _cashflow_import_jobs_lock:
            _cashflow_import_jobs[job_id] = {
                "id": job_id,
                "status": "queued",
                "created_ts": time.time(),
                "created_at": now_iso,
                "updated_at": now_iso,
                "rows_estimate": max(0, raw_bytes.count(b"\n") - 1),
                "month": month,
                "filename": filename,
                "result": None,
                "error": None,
            }

        def _run_job() -> None:
            with _cashflow_import_jobs_lock:
                row = _cashflow_import_jobs.get(job_id)
                if not row:
                    return
                row["status"] = "running"
                row["updated_at"] = datetime.now(timezone.utc).isoformat()

            try:
                result = _execute_cashflow_import(
                    month=month,
                    filename=filename,
                    raw_bytes=raw_bytes,
                    force=force,
                    merge_strategy=merge_strategy,
                )
                with _cashflow_import_jobs_lock:
                    row = _cashflow_import_jobs.get(job_id)
                    if not row:
                        return
                    row["status"] = "done"
                    row["result"] = result
                    row["updated_at"] = datetime.now(timezone.utc).isoformat()
            except Exception as exc:
                with _cashflow_import_jobs_lock:
                    row = _cashflow_import_jobs.get(job_id)
                    if not row:
                        return
                    row["status"] = "failed"
                    row["error"] = str(exc)
                    row["updated_at"] = datetime.now(timezone.utc).isoformat()

        _cashflow_import_executor.submit(_run_job)
        return job_id

    # ── Page ────────────────────────────────────────────────

    def _finance_flags_payload() -> dict[str, bool]:
        defaults: dict[str, bool] = {
            "a11yEnhancements": True,
            "keyboardShortcuts": True,
            "sectionQuickNav": True,
            "featureFlagsUI": True,
            "perfTelemetry": True,
            "lazyRebalanceLoad": True,
            "secondaryPanelCache": True,
        }
        raw = app.config.get("FINANCE_FEATURE_FLAGS")
        if raw is None:
            return defaults

        parsed: dict | None = None
        if isinstance(raw, dict):
            parsed = raw
        elif isinstance(raw, str) and raw.strip():
            try:
                parsed = json.loads(raw)
            except Exception:
                parsed = None

        if not parsed:
            return defaults

        merged = dict(defaults)
        for key in defaults:
            if key in parsed:
                merged[key] = bool(parsed[key])
        return merged

    @app.get("/finance")
    def finance_page():
        return render_template(
            "finance.html",
            finance_flags=_finance_flags_payload(),
        )

    @app.get("/finance/registros")
    def finance_records_page():
        return render_template("finance_records.html")

    # ── Summary ─────────────────────────────────────────────

    @app.get("/api/finance/summary")
    @limiter.limit("30/minute")
    def finance_summary():
        cached = cache.get("finance:summary")
        if cached:
            return jsonify(cached)
        summary = repo.get_fin_summary()
        currency = repo.list_currency_rates()
        summary["currency_rates"] = currency
        cache.set("finance:summary", summary, FINANCE_CACHE_TTLS["summary"])
        return jsonify(summary)

    # ── Finance Settings (keys & providers) ───────────────

    FINANCE_SETTINGS_SCHEMA: dict[str, dict] = {
        "brapi_token": {
            "type": "str", "max_len": 300, "default": "", "secret": True,
        },
        "brapi_monthly_limit": {
            "type": "int", "min": 100, "max": 500000, "default": "15000",
        },
        "brapi_max_calls_per_request": {
            "type": "int", "min": 0, "max": 200, "default": "2",
        },
        "brapi_reserve_pct": {
            "type": "int", "min": 0, "max": 50, "default": "15",
        },
        "finance_api_key": {
            "type": "str", "max_len": 300, "default": "", "secret": True,
        },
        "ai_local_enabled": {"type": "bool", "default": "1"},
        "ai_local_url": {
            "type": "url", "max_len": 300,
            "default": "http://llamacpp:8080",
        },
        "ai_local_model": {
            "type": "str", "max_len": 200,
            "default": "qwen2.5:7b-instruct",
        },
        "ai_local_timeout_seconds": {
            "type": "int", "min": 5, "max": 300, "default": "45",
        },
        "currency_api_url": {
            "type": "url", "max_len": 500,
            "default": "https://economia.awesomeapi.com.br/last/USD-BRL,EUR-BRL,BTC-BRL",
        },
        "currency_update_minutes": {
            "type": "int", "min": 1, "max": 1440, "default": "15",
        },
    }

    def _bool_to_01(value: object) -> str:
        if isinstance(value, bool):
            return "1" if value else "0"
        txt = str(value).strip().lower()
        return "1" if txt in ("1", "true", "yes", "on") else "0"

    def _validate_fin_setting(key: str, value: object) -> tuple[bool, str, str]:
        schema = FINANCE_SETTINGS_SCHEMA.get(key)
        if not schema:
            return False, "", f"Unknown setting: {key}"

        stype = schema["type"]
        raw = str(value).strip()

        if stype == "int":
            try:
                v = int(raw)
                if v < schema.get("min", -9999999) or v > schema.get("max", 9999999):
                    return False, "", f"{key}: valor fora do intervalo"
                return True, str(v), ""
            except ValueError:
                return False, "", f"{key}: deve ser número inteiro"

        if stype == "bool":
            if isinstance(value, bool):
                return True, "1" if value else "0", ""
            if raw.lower() in ("0", "1", "true", "false", "yes", "no", "on", "off"):
                return True, _bool_to_01(raw), ""
            return False, "", f"{key}: deve ser booleano"

        if stype == "url":
            if len(raw) > schema.get("max_len", 500):
                return False, "", f"{key}: URL muito longa"
            return True, raw, ""

        if stype == "str":
            if len(raw) > schema.get("max_len", 500):
                return False, "", f"{key}: texto muito longo"
            return True, raw, ""

        return True, raw, ""

    def _is_finite_number(value: object) -> bool:
        if not isinstance(value, (int, float, str)):
            return False
        try:
            n = float(value)
        except (TypeError, ValueError):
            return False
        return math.isfinite(n)

    def _as_float(value: object, default: float = 0.0) -> float:
        if not isinstance(value, (int, float, str)):
            return default
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    def _normalize_tags(value: object, *, max_items: int = 12) -> list[str]:
        items: list[str] = []
        if isinstance(value, list):
            items = [str(x) for x in value]
        elif isinstance(value, str):
            items = [x.strip() for x in value.split(",")]
        else:
            return []

        out: list[str] = []
        seen: set[str] = set()
        for raw in items:
            tag = sanitize_text(str(raw or "").strip().lower(), 30)
            if not tag or tag in seen:
                continue
            seen.add(tag)
            out.append(tag)
            if len(out) >= max_items:
                break
        return out

    def _track_api_provider_usage(provider: str, success: bool, endpoint: str = "") -> None:
        try:
            today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            base_key = f"api_usage:{today}:{endpoint}:{provider}"
            total = int(repo.get_setting(base_key, "0") or 0)
            status_key = f"{base_key}:{'ok' if success else 'err'}"
            status = int(repo.get_setting(status_key, "0") or 0)
            repo.set_setting(base_key, str(total + 1))
            repo.set_setting(status_key, str(status + 1))
            cache.delete(f"finance:provider-metrics:{today}")
        except Exception:
            pass

    def _track_api_provider_latency(
        provider: str,
        endpoint: str,
        latency_ms: float,
    ) -> None:
        try:
            if not math.isfinite(float(latency_ms)):
                return
            today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            base_key = f"api_usage:{today}:{endpoint}:{provider}"
            sum_key = f"{base_key}:latency_sum_ms"
            count_key = f"{base_key}:latency_count"
            cur_sum = float(repo.get_setting(sum_key, "0") or 0)
            cur_count = int(repo.get_setting(count_key, "0") or 0)
            repo.set_setting(sum_key, f"{cur_sum + float(latency_ms):.3f}")
            repo.set_setting(count_key, str(cur_count + 1))
            cache.delete(f"finance:provider-metrics:{today}")
        except Exception:
            pass

    def _track_api_provider_fallback(provider: str, endpoint: str = "") -> None:
        try:
            today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            base_key = f"api_usage:{today}:{endpoint}:{provider}"
            fallback_key = f"{base_key}:fallback"
            cur = int(repo.get_setting(fallback_key, "0") or 0)
            repo.set_setting(fallback_key, str(cur + 1))
            cache.delete(f"finance:provider-metrics:{today}")
        except Exception:
            pass

    def _provider_day_metrics(day_key: str) -> dict[str, dict]:
        cache_key = f"finance:provider-metrics:{day_key}"
        cached: dict[str, dict] | None = cache.get(cache_key)
        if cached is not None:
            return cached

        settings = repo.get_all_settings()
        prefix = f"api_usage:{day_key}:"
        providers: dict[str, dict] = {}

        for key, value in settings.items():
            if not key.startswith(prefix):
                continue
            parts = key.split(":")
            if len(parts) < 4:
                continue
            endpoint = str(parts[2] or "")
            provider = str(parts[3] or "")
            suffix = parts[4] if len(parts) >= 5 else ""
            if not endpoint or not provider:
                continue

            p = providers.setdefault(
                provider,
                {
                    "provider": provider,
                    "total": 0,
                    "ok": 0,
                    "err": 0,
                    "fallback": 0,
                    "latency_sum_ms": 0.0,
                    "latency_count": 0,
                    "endpoints": {},
                },
            )
            ep = p["endpoints"].setdefault(
                endpoint,
                {
                    "total": 0,
                    "ok": 0,
                    "err": 0,
                    "fallback": 0,
                    "latency_sum_ms": 0.0,
                    "latency_count": 0,
                },
            )

            if suffix == "ok":
                n = int(value or 0)
                p["ok"] += n
                ep["ok"] += n
            elif suffix == "err":
                n = int(value or 0)
                p["err"] += n
                ep["err"] += n
            elif suffix == "fallback":
                n = int(value or 0)
                p["fallback"] += n
                ep["fallback"] += n
            elif suffix == "latency_sum_ms":
                n = float(value or 0)
                p["latency_sum_ms"] += n
                ep["latency_sum_ms"] += n
            elif suffix == "latency_count":
                n = int(value or 0)
                p["latency_count"] += n
                ep["latency_count"] += n
            elif suffix == "":
                n = int(value or 0)
                p["total"] += n
                ep["total"] += n

        for provider_metrics in providers.values():
            total = int(provider_metrics.get("total") or 0)
            ok = int(provider_metrics.get("ok") or 0)
            fb = int(provider_metrics.get("fallback") or 0)
            lsum = float(provider_metrics.get("latency_sum_ms") or 0)
            lcount = int(provider_metrics.get("latency_count") or 0)
            provider_metrics["success_rate"] = round((ok / total), 4) if total > 0 else None
            provider_metrics["fallback_rate"] = round((fb / total), 4) if total > 0 else None
            provider_metrics["avg_latency_ms"] = round((lsum / lcount), 2) if lcount > 0 else None
            provider_metrics.pop("latency_sum_ms", None)

            for endpoint_metrics in provider_metrics["endpoints"].values():
                etotal = int(endpoint_metrics.get("total") or 0)
                eok = int(endpoint_metrics.get("ok") or 0)
                efb = int(endpoint_metrics.get("fallback") or 0)
                elsum = float(endpoint_metrics.get("latency_sum_ms") or 0)
                elcount = int(endpoint_metrics.get("latency_count") or 0)
                endpoint_metrics["success_rate"] = (
                    round((eok / etotal), 4)
                    if etotal > 0
                    else None
                )
                endpoint_metrics["fallback_rate"] = (
                    round((efb / etotal), 4)
                    if etotal > 0
                    else None
                )
                endpoint_metrics["avg_latency_ms"] = (
                    round((elsum / elcount), 2)
                    if elcount > 0
                    else None
                )
                endpoint_metrics.pop("latency_sum_ms", None)

        cache.set(cache_key, providers, 60)
        return providers

    def _audit(
        action: str,
        target_type: str,
        target_id: int | None,
        payload: dict | None = None,
    ) -> None:
        try:
            forwarded_for = str(
                request.headers.get("X-Forwarded-For", ""),
            ).split(",")[0].strip()
            meta = {
                "ip": forwarded_for or str(request.remote_addr or ""),
                "ua": str(request.user_agent.string or "")[:180],
                "path": str(request.path or ""),
                "method": str(request.method or ""),
            }
            merged_payload = dict(payload or {})
            merged_payload["_meta"] = meta
            repo.add_fin_audit_log(
                action,
                target_type,
                target_id,
                merged_payload,
            )
            _invalidate_cache_prefixes("finance:audit:")
        except Exception as exc:
            logger.warning("finance audit failed: %s", exc)

    # ── Register Cashflow Blueprint ─────────────────────────────────────────
    
    _cashflow_helpers = {
        '_audit': _audit,
        '_invalidate_cashflow_cache': _invalidate_cashflow_cache,
        '_as_float': _as_float,
        'FINANCE_CACHE_TTLS': FINANCE_CACHE_TTLS,
    }
    register_cashflow_routes(app, limiter, repo, cache, logger, _cashflow_helpers)

    _watchlist_helpers = {
        '_audit': _audit,
        '_invalidate_financial_state_cache': _invalidate_financial_state_cache,
        '_is_finite_number': _is_finite_number,
    }
    register_watchlist_routes(app, limiter, repo, cache, logger, _watchlist_helpers)

    _assets_helpers = {
        '_audit': _audit,
        '_invalidate_financial_state_cache': _invalidate_financial_state_cache,
        '_is_finite_number': _is_finite_number,
        '_recalc_portfolio': _recalc_portfolio,
        'finance_asset_history_alt': lambda asset_id: finance_asset_history_alt(asset_id),
    }
    _admin_helpers = {
        'FINANCE_SETTINGS_SCHEMA': FINANCE_SETTINGS_SCHEMA,
        '_bool_to_01': _bool_to_01,
        '_validate_fin_setting': _validate_fin_setting,
        '_invalidate_financial_state_cache': _invalidate_financial_state_cache,
        '_provider_day_metrics': _provider_day_metrics,
    }
    _goals_helpers = {
        'FINANCE_CACHE_TTLS': FINANCE_CACHE_TTLS,
        '_audit': _audit,
        '_invalidate_cache_prefixes': _invalidate_cache_prefixes,
        '_is_finite_number': _is_finite_number,
    }
    _credit_card_helpers = {
        '_audit': _audit,
        '_is_finite_number': _is_finite_number,
    }
    _debt_helpers = {
        'FINANCE_CACHE_TTLS': FINANCE_CACHE_TTLS,
    }
    _account_helpers = {
        '_audit': _audit,
        '_is_finite_number': _is_finite_number,
    }
    _analytics_helpers = {
        'FINANCE_CACHE_TTLS': FINANCE_CACHE_TTLS,
    }
    _anomalies_helpers = {
        '_fmt_brl': _fmt_brl,
    }
    _insights_helpers = {
        '_fmt_brl': _fmt_brl,
    }
    _maintenance_helpers = {
        'CLEANUP_CONFIRM_TOKEN': CLEANUP_CONFIRM_TOKEN,
        'CLEANUP_COOLDOWN_SECONDS': CLEANUP_COOLDOWN_SECONDS,
        'CLEANUP_IDEMPOTENCY_TTL': CLEANUP_IDEMPOTENCY_TTL,
        '_audit': _audit,
        '_invalidate_financial_state_cache': _invalidate_financial_state_cache,
        '_recalc_portfolio': _recalc_portfolio,
    }
    _monthly_comparison_helpers = {
        'FINANCE_CACHE_TTLS': FINANCE_CACHE_TTLS,
    }
    _dividends_helpers = {
        'FINANCE_CACHE_TTLS': FINANCE_CACHE_TTLS,
        '_invalidate_financial_state_cache': _invalidate_financial_state_cache,
    }
    _performance_helpers = {
        'FINANCE_CACHE_TTLS': FINANCE_CACHE_TTLS,
        '_track_api_provider_usage': _track_api_provider_usage,
        '_track_api_provider_latency': _track_api_provider_latency,
    }
    _investment_planning_helpers = {
        'FINANCE_CACHE_TTLS': FINANCE_CACHE_TTLS,
        '_invalidate_cache_prefixes': _invalidate_cache_prefixes,
    }
    register_analytics_routes(app, limiter, repo, cache, logger, _analytics_helpers)
    register_anomaly_routes(app, limiter, repo, cache, logger, _anomalies_helpers)
    register_admin_routes(app, limiter, repo, cache, logger, _admin_helpers)
    register_account_routes(app, limiter, repo, cache, logger, _account_helpers)
    register_credit_card_routes(app, limiter, repo, cache, logger, _credit_card_helpers)
    register_debt_routes(app, limiter, repo, cache, logger, _debt_helpers)
    register_dividends_routes(app, limiter, repo, cache, logger, _dividends_helpers)
    register_goals_routes(app, limiter, repo, cache, logger, _goals_helpers)
    register_insights_routes(app, limiter, repo, cache, logger, _insights_helpers)
    register_investment_planning_routes(app, limiter, repo, cache, logger, _investment_planning_helpers)
    register_maintenance_routes(app, limiter, repo, cache, logger, _maintenance_helpers)
    register_monthly_comparison_routes(app, limiter, repo, cache, logger, _monthly_comparison_helpers)
    register_performance_routes(app, limiter, repo, cache, logger, _performance_helpers)
    register_assets_routes(app, limiter, repo, cache, logger, _assets_helpers)
    register_security_routes(app, limiter, repo, cache, logger)

    # ── Assets CRUD ─────────────────────────────────────────

    # ── Split entry ────────────────────────────────────────

    @app.put("/api/finance/cashflow/saved-filters/<int:filter_id>/favorite")
    @limiter.limit("30/minute")
    @require_finance_key
    def finance_cashflow_saved_filters_toggle_favorite(filter_id: int):
        """Toggle favorite status on a saved filter."""
        rows = repo.list_saved_filters()
        target = next((r for r in rows if int(r.get("id") or 0) == filter_id), None)
        if not target:
            return jsonify({"error": "Filtro não encontrado"}), 404
        if not str(target.get("name") or "").startswith("cashflow:"):
            return jsonify({"error": "Filtro não pertence ao cashflow"}), 400
        
        success = repo.toggle_saved_filter_favorite(filter_id)
        is_favorite = not bool(target.get("is_favorite"))
        return jsonify({
            "ok": success,
            "id": filter_id,
            "is_favorite": is_favorite,
        })

    @app.post("/api/finance/cashflow/saved-filters/<int:filter_id>/apply")
    @limiter.limit("30/minute")
    def finance_cashflow_saved_filters_apply(filter_id: int):
        """Apply a saved filter and track usage."""
        rows = repo.list_saved_filters()
        target = next((r for r in rows if int(r.get("id") or 0) == filter_id), None)
        if not target:
            return jsonify({"error": "Filtro não encontrado"}), 404
        
        # Track usage
        repo.track_filter_usage(filter_id)
        
        return jsonify({
            "ok": True,
            "filter": target.get("filter") if isinstance(target.get("filter"), dict) else {},
            "usage_count": int(target.get("use_count") or 0) + 1,
        })

    @app.get("/api/finance/cashflow/filters/templates")
    @limiter.limit("30/minute")
    def finance_cashflow_filters_templates():
        """Get predefined filter templates."""
        templates = repo.get_saved_filter_templates()
        payload = []
        for row in templates:
            name = str(row.get("name") or "")
            if not name.startswith("cashflow:"):
                continue
            payload.append({
                "id": int(row.get("id") or 0),
                "name": name.replace("cashflow:", "", 1),
                "description": row.get("description") or "",
                "filter": row.get("filter") if isinstance(row.get("filter"), dict) else {},
            })
        return jsonify(payload)

    @app.get("/api/finance/cashflow/data-quality")
    @limiter.limit("30/minute")
    def finance_cashflow_data_quality():
        month = sanitize_text(str(request.args.get("month", "")), 7).strip()
        if not month:
            month = datetime.now(timezone.utc).strftime("%Y-%m")
        if not re.match(r"^\d{4}-\d{2}$", month):
            return jsonify({"error": "month inválido (use YYYY-MM)"}), 400

        rows = repo.list_fin_cashflow_entries(month=month, limit=5000)
        missing_category = 0
        missing_description = 0
        future_dates = 0
        non_positive_amount = 0
        duplicate_map: dict[str, list[dict]] = {}
        expense_amounts: list[float] = []
        today = datetime.now(timezone.utc).date()

        for row in rows:
            category = str(row.get("category") or "").strip()
            description = str(row.get("description") or "").strip()
            entry_date = str(row.get("entry_date") or "")[:10]
            amount = float(row.get("amount") or 0)
            entry_type = str(row.get("entry_type") or "expense").strip().lower()

            if not category:
                missing_category += 1
            if not description:
                missing_description += 1
            if amount <= 0:
                non_positive_amount += 1
            if entry_type == "expense" and amount > 0:
                expense_amounts.append(amount)

            if re.match(r"^\d{4}-\d{2}-\d{2}$", entry_date):
                try:
                    dt = datetime.strptime(entry_date, "%Y-%m-%d").date()
                    if dt > today:
                        future_dates += 1
                except ValueError:
                    pass

            dup_key = cashflow_dedupe_hash(
                entry_type,
                amount,
                entry_date,
                description,
            )
            duplicate_map.setdefault(dup_key, []).append(row)

        duplicate_groups = [items for items in duplicate_map.values() if len(items) > 1]
        duplicate_rows = sum(max(0, len(items) - 1) for items in duplicate_groups)

        fuzzy_duplicate_pairs = 0
        fuzzy_duplicate_samples: list[dict[str, Any]] = []
        processed_pairs: set[tuple[int, int]] = set()
        rows_by_type = {
            "income": [r for r in rows if str(r.get("entry_type") or "").strip().lower() == "income"],
            "expense": [r for r in rows if str(r.get("entry_type") or "").strip().lower() == "expense"],
        }
        for entry_type, pool in rows_by_type.items():
            for row in pool:
                row_id = int(row.get("id") or 0)
                if not row_id:
                    continue
                candidate = find_potential_cashflow_duplicate(
                    existing_entries=pool,
                    entry_type=entry_type,
                    amount=float(row.get("amount") or 0),
                    entry_date=str(row.get("entry_date") or "")[:10],
                    description=str(row.get("description") or ""),
                )
                if not candidate:
                    continue
                cand_id = int(candidate.get("id") or 0)
                if not cand_id or cand_id == row_id:
                    continue
                pair = tuple(sorted((row_id, cand_id)))
                pair = cast(tuple[int, int], pair)
                if pair in processed_pairs:
                    continue
                processed_pairs.add(pair)
                fuzzy_duplicate_pairs += 1
                if len(fuzzy_duplicate_samples) < 6:
                    fuzzy_duplicate_samples.append({
                        "ids": [pair[0], pair[1]],
                        "amount": round(float(row.get("amount") or 0), 2),
                        "entry_date": str(row.get("entry_date") or "")[:10],
                        "description": str(row.get("description") or ""),
                        "score": float(candidate.get("score") or 0),
                        "confidence": str(candidate.get("confidence") or "low"),
                    })

        outlier_count = 0
        outlier_samples: list[dict] = []
        if len(expense_amounts) >= 6:
            ordered = sorted(expense_amounts)
            median = ordered[len(ordered) // 2]
            threshold = max(500.0, median * 3.0)
            for row in rows:
                if str(row.get("entry_type") or "").strip().lower() != "expense":
                    continue
                amount = float(row.get("amount") or 0)
                if amount > threshold:
                    outlier_count += 1
                    if len(outlier_samples) < 5:
                        outlier_samples.append({
                            "id": int(row.get("id") or 0),
                            "amount": round(amount, 2),
                            "category": str(row.get("category") or ""),
                            "description": str(row.get("description") or ""),
                        })

        total_rows = len(rows)
        score = 100
        score -= min(30, missing_category * 2)
        score -= min(20, missing_description)
        score -= min(15, duplicate_rows * 3)
        score -= min(10, fuzzy_duplicate_pairs * 2)
        score -= min(15, future_dates * 5)
        score -= min(10, non_positive_amount * 5)
        score -= min(10, outlier_count * 2)
        score = max(0, int(score))

        duplicate_samples: list[dict] = []
        for group in duplicate_groups[:4]:
            first = group[0]
            duplicate_samples.append({
                "ids": [int(i.get("id") or 0) for i in group[:4]],
                "entry_date": str(first.get("entry_date") or "")[:10],
                "amount": round(float(first.get("amount") or 0), 2),
                "description": str(first.get("description") or ""),
                "count": len(group),
            })

        issues = [
            {
                "code": "missing_category",
                "severity": "high" if missing_category > 0 else "ok",
                "count": missing_category,
                "label": "Lançamentos sem categoria",
            },
            {
                "code": "missing_description",
                "severity": "medium" if missing_description > 0 else "ok",
                "count": missing_description,
                "label": "Lançamentos sem descrição",
            },
            {
                "code": "duplicates",
                "severity": "high" if duplicate_rows > 0 else "ok",
                "count": duplicate_rows,
                "label": "Possíveis duplicatas",
                "samples": duplicate_samples,
            },
            {
                "code": "near_duplicates",
                "severity": "medium" if fuzzy_duplicate_pairs > 0 else "ok",
                "count": fuzzy_duplicate_pairs,
                "label": "Possíveis duplicatas aproximadas",
                "samples": fuzzy_duplicate_samples,
            },
            {
                "code": "future_dates",
                "severity": "medium" if future_dates > 0 else "ok",
                "count": future_dates,
                "label": "Lançamentos com data futura",
            },
            {
                "code": "outliers",
                "severity": "medium" if outlier_count > 0 else "ok",
                "count": outlier_count,
                "label": "Despesas fora do padrão",
                "samples": outlier_samples,
            },
        ]

        # Add quality alerts (NEW)
        alerts_payload = evaluate_data_quality_alerts(
            score=score,
            issues=issues,
        )

        return jsonify({
            "month": month,
            "score": score,
            "total_rows": total_rows,
            "issues": issues,
            "suggestions": [
                "Preencha categoria e descrição nos lançamentos incompletos.",
                "Revise duplicatas antes de fechar o mês.",
                "Valide despesas fora do padrão para evitar classificação incorreta.",
            ],
            "quality_alerts": alerts_payload,
        })

    @app.get("/api/finance/cashflow/bulk/dedup-stats")
    @limiter.limit("30/minute")
    def finance_cashflow_dedup_stats():
        """Get dedup cache statistics with per-operation metrics."""
        stats = _dedupe_cache.stats()
        return jsonify({
            "ok": True,
            "cache_backend": (
                "redis" if hasattr(cache, "redis")
                else "memory"
            ),
            "cache_stats": stats,
            "recommendation": (
                "Cache working well" if stats.get("hit_rate", 0) >= 0.5
                else "Low hit rate: consider increasing TTL or request volume"
            ),
        })

    @app.post("/api/finance/cashflow/bulk/dedup-reset")
    @limiter.limit("5/minute")
    @require_finance_key
    def finance_cashflow_dedup_reset():
        """Clear dedup cache (admin only)."""
        _dedupe_cache.clear()
        _audit("dedup_cache_reset", "cashflow", None, {
            "reason": "Manual reset via admin endpoint",
        })
        return jsonify({"ok": True, "message": "Dedup cache cleared"})

    @app.get("/api/finance/savings-suggestions")
    @limiter.limit("30/minute")
    def finance_savings_suggestions():
        month = sanitize_text(str(request.args.get("month", "")), 7).strip()
        if month and not re.match(r"^\d{4}-\d{2}$", month):
            return jsonify({"error": "month invalid (YYYY-MM)"}), 400
        target_month = month or datetime.now(timezone.utc).strftime("%Y-%m")
        cache_key = f"finance:savings-suggestions:{target_month}"
        cached = cache.get(cache_key)
        if cached:
            return jsonify(cached)
        try:
            payload = repo.get_fin_savings_suggestions(target_month)
            cache.set(cache_key, payload, FINANCE_CACHE_TTLS["cashflow_analytics"])
            return jsonify(payload)
        except Exception as ex:
            return jsonify({"error": str(ex)}), 500

    @app.patch("/api/finance/cashflow/<int:entry_id>")
    @require_finance_key
    def finance_update_cashflow_inline(entry_id: int):
        data = request.get_json(silent=True) or {}
        allowed_fields = {"category", "description", "amount", "account_id", "credit_card_id"}
        updates = {k: v for k, v in data.items() if k in allowed_fields}
        if not updates:
            return jsonify({"error": "no updates"}), 400
        try:
            result = repo.update_fin_cashflow_entry(entry_id, updates)
            if result:
                _invalidate_cashflow_cache()
                return jsonify({"status": "ok"}), 200
            return jsonify({"error": "not found"}), 404
        except Exception as ex:
            return jsonify({"error": str(ex)}), 500

    @app.get("/api/finance/report/pdf")
    @limiter.limit("10/minute")
    def finance_report_pdf():
        month = sanitize_text(str(request.args.get("month", "")), 7).strip()
        if not month or not re.match(r"^\d{4}-\d{2}$", month):
            return jsonify({"error": "month invalid"}), 400
        theme = sanitize_text(str(request.args.get("theme", "light")), 10).strip().lower()
        if theme not in ("light", "dark"):
            theme = "light"
        try:
            summary = repo.get_fin_cashflow_summary(months=18)
            analytics = repo.get_fin_cashflow_analytics(month=month)
            totals = analytics.get("totals", {})
            month_rows = [
                row for row in (summary.get("monthly") or []) if str(row.get("month") or "") == month
            ]
            month_row = month_rows[0] if month_rows else {}
            lines = [
                "Relatorio Financeiro",
                f"Mes de referencia: {month}",
                f"Tema: {theme}",
                f"Gerado em: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
                "",
                "Resumo do mes:",
                f"- Receitas: R$ {float(totals.get('income') or 0):,.2f}",
                f"- Despesas: R$ {float(totals.get('expense') or 0):,.2f}",
                f"- Saldo: R$ {float(totals.get('balance') or 0):,.2f}",
                f"- Taxa de poupanca: {float(totals.get('savings_rate_pct') or 0):.2f}%",
                "",
                "Top despesas:",
            ]
            for row in (analytics.get("top_expenses") or [])[:5]:
                lines.append(f"- {row.get('category')}: R$ {float(row.get('amount') or 0):,.2f}")
            lines.extend([
                "",
                "Comparativo serie mensal:",
                f"- Receita do mes: R$ {float(month_row.get('income') or 0):,.2f}",
                f"- Despesa do mes: R$ {float(month_row.get('expense') or 0):,.2f}",
                f"- Saldo do mes: R$ {float(month_row.get('balance') or 0):,.2f}",
            ])
            pdf_bytes = _build_simple_pdf(lines, theme=theme)
            return pdf_bytes, 200, {
                "Content-Type": "application/pdf",
                "Content-Disposition": f"attachment; filename=relatorio-{month}.pdf",
            }
        except Exception as ex:
            return jsonify({"error": str(ex)}), 500

    @app.post("/api/finance/cashflow/budget/template")
    @limiter.limit("20/minute")
    @require_finance_key
    def finance_cashflow_budget_template():
        """Aplica metodologia de orçamento pessoal (50/30/20, envelope, base-zero)."""
        body = request.get_json(silent=True) or {}
        method = sanitize_text(str(body.get("method", "50_30_20")), 20).strip()
        month = sanitize_text(str(body.get("month", "")), 7).strip()
        if not month:
            month = datetime.now(timezone.utc).strftime("%Y-%m")
        if not re.match(r"^\d{4}-\d{2}$", month):
            return jsonify({"error": "month inválido (use YYYY-MM)"}), 400
        try:
            income = float(body.get("income", 0))
        except (TypeError, ValueError):
            income = 0.0
        if not _is_finite_number(income) or income <= 0:
            return jsonify({"error": "income deve ser um valor positivo"}), 400
        apply = bool(body.get("apply", False))

        VALID_METHODS = {"50_30_20", "envelope", "zero_based"}
        if method not in VALID_METHODS:
            return jsonify({"error": f"method inválido. Use: {', '.join(sorted(VALID_METHODS))}"}), 400

        # Category name → (group, percentage of income)
        TEMPLATES: dict[str, dict] = {
            "50_30_20": {
                "label": "50/30/20 — Necessidades • Desejos • Poupança",
                "description": (
                    "50% da renda líquida para necessidades básicas, "
                    "30% para desejos pessoais e 20% para poupança/investimentos."
                ),
                "groups": {"Necessidades": 0.50, "Desejos": 0.30, "Poupança": 0.20},
                "categories": {
                    "Moradia": 0.25, "Alimentação": 0.10, "Transporte": 0.08,
                    "Saúde": 0.05, "Contas": 0.02,
                    "Restaurante": 0.08, "Lazer": 0.08, "Compras": 0.07,
                    "Assinaturas": 0.04, "Viagem": 0.03,
                    "Investimentos": 0.15, "Reserva": 0.05,
                },
            },
            "envelope": {
                "label": "Orçamento por Envelopes",
                "description": (
                    "Aloque valores fixos para cada 'envelope' (categoria) "
                    "até esgotar o total da renda disponível."
                ),
                "groups": {"Total": 1.0},
                "categories": {
                    "Moradia": 0.28, "Alimentação": 0.12, "Transporte": 0.10,
                    "Saúde": 0.08, "Lazer": 0.10, "Restaurante": 0.08,
                    "Compras": 0.07, "Assinaturas": 0.05, "Investimentos": 0.12,
                },
            },
            "zero_based": {
                "label": "Orçamento Base Zero",
                "description": (
                    "Cada real da renda é alocado explicitamente — "
                    "receita menos todas as despesas e investimentos deve ser R$0."
                ),
                "groups": {"Total": 1.0},
                "categories": {
                    "Moradia": 0.25, "Alimentação": 0.10, "Transporte": 0.08,
                    "Saúde": 0.05, "Contas": 0.03, "Restaurante": 0.07,
                    "Lazer": 0.07, "Compras": 0.06, "Assinaturas": 0.04,
                    "Viagem": 0.03, "Investimentos": 0.15, "Reserva": 0.04,
                    "Outros": 0.03,
                },
            },
        }

        tmpl = TEMPLATES[method]
        budget: dict[str, float] = {
            cat: round(income * pct, 2)
            for cat, pct in tmpl["categories"].items()
        }
        groups: dict[str, float] = {
            g: round(income * pct, 2)
            for g, pct in tmpl["groups"].items()
        }

        if apply:
            repo.set_fin_cashflow_budget(month, budget)
            _audit("update", "cashflow_budget", None, {
                "method": method, "month": month,
                "income": income, "categories": sorted(budget.keys()),
            })

        return jsonify({
            "ok": True,
            "method": method,
            "label": tmpl["label"],
            "description": tmpl["description"],
            "income": income,
            "month": month,
            "applied": apply,
            "budget": budget,
            "groups": groups,
        })

    @app.get("/api/finance/cashflow/budget/check")
    @limiter.limit("60/minute")
    def finance_cashflow_budget_check():
        """Verifica se um gasto ultrapassaria o limite orçado para a categoria."""
        category = sanitize_text(str(request.args.get("category", "")), 60).strip()
        month = sanitize_text(str(request.args.get("month", "")), 7).strip()
        if not month:
            month = datetime.now(timezone.utc).strftime("%Y-%m")
        if not re.match(r"^\d{4}-\d{2}$", month):
            return jsonify({"error": "month inválido (use YYYY-MM)"}), 400
        if not category:
            return jsonify({"error": "category obrigatória"}), 400
        try:
            amount = float(request.args.get("amount", 0))
        except (TypeError, ValueError):
            amount = 0.0
        if not _is_finite_number(amount) or amount < 0:
            amount = 0.0

        budget = repo.get_fin_cashflow_budget(month)
        limit = budget.get(category)
        if limit is None:
            return jsonify({
                "category": category, "month": month,
                "has_budget": False, "limit": None,
                "spent": 0, "remaining": None, "over_budget": False,
            })

        analytics = repo.get_fin_cashflow_analytics(month=month)
        expense_cats = analytics.get("categories", {}).get("expense", [])
        spent = next(
            (float(r["amount"]) for r in expense_cats if r["category"] == category),
            0.0,
        )
        would_be = spent + amount
        return jsonify({
            "category": category,
            "month": month,
            "has_budget": True,
            "limit": round(limit, 2),
            "spent": round(spent, 2),
            "would_be_spent": round(would_be, 2),
            "remaining": round(limit - spent, 2),
            "over_budget": would_be > limit,
            "usage_pct": round(would_be / limit * 100 if limit > 0 else 0, 2),
        })

    @app.get("/api/finance/cashflow/budget/alerts")
    @limiter.limit("30/minute")
    def finance_cashflow_budget_alerts():
        month = sanitize_text(str(request.args.get("month", "")), 7).strip()
        if not month:
            month = datetime.now(timezone.utc).strftime("%Y-%m")
        if not re.match(r"^\d{4}-\d{2}$", month):
            return jsonify({"error": "month inválido (use YYYY-MM)"}), 400
        try:
            threshold = float(request.args.get("threshold", 80))
        except (TypeError, ValueError):
            threshold = 80.0
        threshold = max(0.0, min(200.0, threshold))

        analytics = repo.get_fin_cashflow_analytics(month=month)
        budget_items = analytics.get("budget", {}).get("items", [])
        alerts = []
        for item in budget_items:
            usage = float(item.get("usage_pct") or 0)
            if item.get("over_budget"):
                status = "over"
            elif usage >= threshold:
                status = "warning"
            else:
                status = "ok"
            alerts.append({
                "category": item["category"],
                "limit": item["limit"],
                "spent": item["spent"],
                "remaining": item["remaining"],
                "usage_pct": item["usage_pct"],
                "status": status,
            })
        alerts_only = [a for a in alerts if a["status"] != "ok"]
        return jsonify({
            "month": month,
            "threshold": threshold,
            "alerts": alerts_only,
            "all": alerts,
        })

    def _load_cashflow_classify_rules() -> list[dict[str, str]]:
        raw_rules = repo.get_setting("cashflow_classify_rules", "[]")
        try:
            parsed = json.loads(raw_rules or "[]")
        except (TypeError, ValueError, json.JSONDecodeError):
            parsed = []

        safe_rules: list[dict[str, str]] = []
        if isinstance(parsed, list):
            for item in parsed:
                if not isinstance(item, dict):
                    continue
                keyword = sanitize_text(str(item.get("keyword") or ""), 100).strip().lower()
                category = sanitize_text(str(item.get("category") or ""), 60).strip()
                if keyword and category:
                    safe_rules.append({"keyword": keyword, "category": category})
        return safe_rules

    def _infer_cashflow_category_from_text(
        description: str,
        rules: list[dict[str, str]] | None = None,
    ) -> str | None:
        normalized = normalize_cashflow_text(description or "")
        if not normalized:
            return None

        if rules is None:
            rules = _load_cashflow_classify_rules()

        for rule in rules:
            keyword = str(rule.get("keyword") or "").strip().lower()
            category = str(rule.get("category") or "").strip()
            if keyword and category and keyword in normalized:
                return category
        return None

    def _infer_cashflow_entry_type_from_text(
        text: str,
        category: str | None = None,
    ) -> str:
        normalized = normalize_cashflow_text(text or "")
        normalized_category = normalize_cashflow_text(category or "")

        income_keywords = (
            "salario",
            "salário",
            "recebido",
            "recebimento",
            "deposito",
            "depósito",
            "pix recebido",
            "credito",
            "crédito",
            "provento",
            "rendimento",
            "bonus",
            "bônus",
            "comissao",
            "comissão",
        )
        expense_keywords = (
            "compra",
            "debito",
            "débito",
            "boleto",
            "pagamento",
            "conta",
            "despesa",
            "mercado",
            "supermercado",
            "restaurante",
            "farmacia",
            "farmácia",
            "cartao credito",
            "cartão crédito",
            "cartao de credito",
            "cartão de crédito",
            "cartao debito",
            "cartão débito",
        )

        if any(keyword in normalized for keyword in expense_keywords):
            return "expense"
        if any(keyword in normalized for keyword in income_keywords):
            return "income"

        if normalized_category in (
            "salario",
            "salário",
            "proventos",
            "rendimentos",
            "receitas",
            "receita",
        ):
            return "income"

        return "expense"

    @app.post("/api/finance/cashflow/auto-classify")
    @limiter.limit("20/minute")
    @require_finance_key
    def finance_cashflow_auto_classify():
        """Auto-classify unclassified or pending entries using keyword rules."""
        body = request.get_json(silent=True) or {}
        month = sanitize_text(str(body.get("month", "")), 7).strip()
        if not month:
            month = datetime.now(timezone.utc).strftime("%Y-%m")
        if not re.match(r"^\d{4}-\d{2}$", month):
            return jsonify({"error": "month inválido (use YYYY-MM)"}), 400

        rules = _load_cashflow_classify_rules()

        entries = repo.list_fin_cashflow_entries(month=month, limit=5000)
        updated = 0
        for entry in entries:
            cat = str(entry.get("category") or "").strip()
            if cat and cat.lower() not in ("", "sem categoria", "outros"):
                continue
            desc = str(entry.get("description") or "")
            inferred = _infer_cashflow_category_from_text(desc, rules)
            if inferred:
                repo.update_fin_cashflow_entry(entry["id"], {"category": inferred})
                updated += 1

        _audit("update", "cashflow_auto_classify", None, {"month": month, "updated": updated})
        return jsonify({"ok": True, "month": month, "updated": updated})

    @app.get("/api/finance/cashflow/classify-rules")
    @limiter.limit("30/minute")
    def finance_cashflow_classify_rules_get():
        rules = _load_cashflow_classify_rules()
        return jsonify({"rules": rules})

    @app.get("/api/finance/cashflow/month-plan")
    @limiter.limit("30/minute")
    def finance_cashflow_month_plan():
        month = sanitize_text(str(request.args.get("month", "")), 7).strip()
        if not month:
            month = datetime.now(timezone.utc).strftime("%Y-%m")
        if not re.match(r"^\d{4}-\d{2}$", month):
            return jsonify({"error": "month inválido (use YYYY-MM)"}), 400

        entries = repo.list_fin_cashflow_entries(month=month, limit=5000)
        analytics = repo.get_fin_cashflow_analytics(month=month)
        budget_rows = analytics.get("budget", {}).get("items", [])
        recurring = repo.list_fin_cashflow_recurring(active_only=True)

        year = int(month[:4])
        month_num = int(month[5:7])
        month_start = datetime(year, month_num, 1).date()
        if month_num == 12:
            month_end = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            month_end = datetime(year, month_num + 1, 1).date() - timedelta(days=1)
        today = datetime.now(timezone.utc).date()

        weeks: list[dict[str, object]] = []
        start_day = 1
        while start_day <= month_end.day:
            end_day = min(start_day + 6, month_end.day)
            ws = datetime(year, month_num, start_day).date()
            we = datetime(year, month_num, end_day).date()
            weeks.append({
                "label": f"{ws.strftime('%d/%m')} - {we.strftime('%d/%m')}",
                "start": ws.isoformat(),
                "end": we.isoformat(),
                "income": 0.0,
                "expense": 0.0,
                "expected_income": 0.0,
                "expected_expense": 0.0,
            })
            start_day += 7

        def _bucket_for(date_obj: Any) -> dict[str, Any] | None:
            for w in weeks:
                if str(w["start"]) <= date_obj.isoformat() <= str(w["end"]):
                    return w
            return None

        for row in entries:
            raw_date = str(row.get("entry_date") or "")[:10]
            if not re.match(r"^\d{4}-\d{2}-\d{2}$", raw_date):
                continue
            d = datetime.strptime(raw_date, "%Y-%m-%d").date()
            bucket = _bucket_for(d)
            if not bucket:
                continue
            amount = float(row.get("amount") or 0)
            if str(row.get("entry_type") or "") == "income":
                bucket["income"] = float(bucket["income"]) + amount
            else:
                bucket["expense"] = float(bucket["expense"]) + amount

        for tpl in recurring:
            try:
                day = int(tpl.get("day_of_month") or 1)
            except (TypeError, ValueError):
                day = 1
            day = max(1, min(month_end.day, day))
            d = datetime(year, month_num, day).date()
            if d < max(today, month_start):
                continue
            bucket = _bucket_for(d)
            if not bucket:
                continue
            amount = float(tpl.get("amount") or 0)
            if str(tpl.get("entry_type") or "") == "income":
                bucket["expected_income"] = float(bucket["expected_income"]) + amount
            else:
                bucket["expected_expense"] = float(bucket["expected_expense"]) + amount

        running_balance = float(analytics.get("totals", {}).get("balance") or 0)
        for bucket in weeks:
            realized = _as_float(bucket.get("income")) - _as_float(bucket.get("expense"))
            expected = _as_float(bucket.get("expected_income")) - _as_float(bucket.get("expected_expense"))
            bucket["net"] = round(realized + expected, 2)
            running_balance += realized + expected
            bucket["projected_balance"] = round(running_balance, 2)
            bucket["income"] = round(_as_float(bucket.get("income")), 2)
            bucket["expense"] = round(_as_float(bucket.get("expense")), 2)
            bucket["expected_income"] = round(_as_float(bucket.get("expected_income")), 2)
            bucket["expected_expense"] = round(_as_float(bucket.get("expected_expense")), 2)

        risks = [
            {
                "category": r.get("category"),
                "usage_pct": r.get("usage_pct"),
                "remaining": r.get("remaining"),
            }
            for r in budget_rows
            if float(r.get("usage_pct") or 0) >= 85
        ]

        return jsonify({
            "month": month,
            "weeks": weeks,
            "risk_flags": risks,
            "starting_balance": round(float(analytics.get("totals", {}).get("balance") or 0), 2),
            "projected_end_balance": round(running_balance, 2),
        })

    @app.put("/api/finance/cashflow/classify-rules")
    @limiter.limit("15/minute")
    @require_finance_key
    def finance_cashflow_classify_rules_put():
        body = request.get_json(silent=True) or {}
        raw_rules = body.get("rules", [])
        if not isinstance(raw_rules, list):
            return jsonify({"error": "rules deve ser uma lista"}), 400
        safe_rules = []
        for r in raw_rules:
            kw = sanitize_text(str(r.get("keyword") or ""), 100).strip().lower()
            cat = sanitize_text(str(r.get("category") or ""), 60).strip()
            if kw and cat:
                safe_rules.append({"keyword": kw, "category": cat})
        repo.set_setting("cashflow_classify_rules", json.dumps(safe_rules))
        _audit("update", "cashflow_classify_rules", None, {"count": len(safe_rules)})
        return jsonify({"ok": True, "rules": safe_rules})

    @app.post("/api/finance/cashflow/scenario")
    @limiter.limit("20/minute")
    def finance_cashflow_scenario():
        """Simula o impacto de reduzir gastos em categorias."""
        body = request.get_json(silent=True) or {}
        month = sanitize_text(str(body.get("month", "")), 7).strip()
        if not month:
            month = datetime.now(timezone.utc).strftime("%Y-%m")
        if not re.match(r"^\d{4}-\d{2}$", month):
            return jsonify({"error": "month inválido (use YYYY-MM)"}), 400

        # adjustments: [{"category": "...", "reduction_pct": 20}]
        raw_adj = body.get("adjustments", [])
        if not isinstance(raw_adj, list):
            return jsonify({"error": "adjustments deve ser uma lista"}), 400

        adjustments: dict[str, float] = {}
        for adj in raw_adj:
            cat = sanitize_text(str(adj.get("category") or ""), 60).strip()
            try:
                pct = float(adj.get("reduction_pct") or 0)
            except (TypeError, ValueError):
                pct = 0.0
            pct = max(-100.0, min(100.0, pct))
            if cat:
                adjustments[cat] = pct

        analytics = repo.get_fin_cashflow_analytics(month=month)
        totals = analytics.get("totals", {})
        current_income = float(totals.get("income") or 0)
        current_expense = float(totals.get("expense") or 0)
        expense_cats = {
            c["category"]: float(c["amount"])
            for c in analytics.get("categories", {}).get("expense", [])
        }

        simulated_expense = 0.0
        scenario_detail = []
        for cat, amount in expense_cats.items():
            reduction_pct = adjustments.get(cat, 0.0)
            simulated = amount * (1 - reduction_pct / 100.0)
            simulated = max(0.0, simulated)
            simulated_expense += simulated
            scenario_detail.append({
                "category": cat,
                "current": round(amount, 2),
                "simulated": round(simulated, 2),
                "saving": round(amount - simulated, 2),
                "reduction_pct": round(reduction_pct, 2),
            })

        simulated_balance = current_income - simulated_expense
        current_balance = current_income - current_expense
        extra_savings = simulated_balance - current_balance
        monthly_saving = round(extra_savings, 2)
        yearly_saving = round(extra_savings * 12, 2)

        return jsonify({
            "month": month,
            "current": {
                "income": round(current_income, 2),
                "expense": round(current_expense, 2),
                "balance": round(current_balance, 2),
            },
            "simulated": {
                "expense": round(simulated_expense, 2),
                "balance": round(simulated_balance, 2),
            },
            "impact": {
                "monthly_saving": monthly_saving,
                "yearly_saving": yearly_saving,
            },
            "detail": scenario_detail,
        })

    @app.post("/api/finance/cashflow/import")
    @limiter.limit("10/minute")
    @require_finance_key
    def finance_cashflow_import():
        """Importa lançamentos de CSV/OFX, com dedupe inteligente e modo assíncrono."""
        month = sanitize_text(str(request.args.get("month", "")), 7).strip()
        if not month:
            month = datetime.now(timezone.utc).strftime("%Y-%m")
        if not re.match(r"^\d{4}-\d{2}$", month):
            return jsonify({"error": "month inválido (use YYYY-MM)"}), 400

        # force=true skips duplicate detection and inserts everything
        force = str(request.args.get("force", "0")).strip().lower() in ("1", "true")
        async_mode = str(request.args.get("async", "0")).strip().lower() in ("1", "true")
        merge_strategy = sanitize_text(str(request.args.get("merge_strategy", "suggest")), 30).strip().lower()
        if merge_strategy not in ("suggest", "append_notes"):
            merge_strategy = "suggest"

        if "file" not in request.files:
            return jsonify({"error": "Nenhum arquivo enviado (campo 'file')"}), 400

        f = request.files["file"]
        filename = (f.filename or "").lower()
        raw_bytes = f.read(2 * 1024 * 1024)  # max 2 MB

        if not (filename.endswith(".csv") or filename.endswith(".ofx") or filename.endswith(".qfx")):
            return jsonify({"error": "Formato não suportado. Use .csv, .ofx ou .qfx"}), 400

        row_estimate = max(0, raw_bytes.count(b"\n") - 1)
        should_async = async_mode or row_estimate >= CASHFLOW_IMPORT_ASYNC_ROW_THRESHOLD
        if should_async:
            job_id = _enqueue_cashflow_import_job(
                month=month,
                filename=filename,
                raw_bytes=raw_bytes,
                force=force,
                merge_strategy=merge_strategy,
            )
            return jsonify({
                "ok": True,
                "async": True,
                "job_id": job_id,
                "status": "queued",
                "rows_estimate": row_estimate,
                "status_url": f"/api/finance/cashflow/import/jobs/{job_id}",
            }), 202

        try:
            result = _execute_cashflow_import(
                month=month,
                filename=filename,
                raw_bytes=raw_bytes,
                force=force,
                merge_strategy=merge_strategy,
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": f"Erro ao processar importação: {exc}"}), 400

        _audit("create", "cashflow_import", None, {
            "filename": filename,
            "imported": int(result.get("imported") or 0),
            "merged": int(result.get("merged") or 0),
            "duplicates_skipped": len(result.get("potential_duplicates") or []),
            "errors": len(result.get("errors") or []),
        })
        return jsonify(result)

    @app.get("/api/finance/cashflow/import/jobs/<string:job_id>")
    @limiter.limit("60/minute")
    @require_finance_key
    def finance_cashflow_import_job_status(job_id: str):
        _cleanup_cashflow_import_jobs()
        with _cashflow_import_jobs_lock:
            row = _cashflow_import_jobs.get(str(job_id or ""))
            if not row:
                return jsonify({"error": "job não encontrado"}), 404
            payload = {
                "ok": True,
                "job_id": row.get("id"),
                "status": row.get("status"),
                "created_at": row.get("created_at"),
                "updated_at": row.get("updated_at"),
                "rows_estimate": row.get("rows_estimate"),
                "month": row.get("month"),
                "filename": row.get("filename"),
                "result": row.get("result"),
                "error": row.get("error"),
            }
        return jsonify(payload)

    def _normalize_ocr_text(raw_text: str) -> str:
        text = str(raw_text or "").replace("\r", "\n")
        text = re.sub(r"[ \t]+", " ", text)
        text = re.sub(r"\n{2,}", "\n", text)
        return text.strip()

    def _sanitize_ocr_manual_text(value: str, max_len: int = 2000) -> str:
        text = html.escape(str(value or "").strip())
        lines = [re.sub(r"\s+", " ", line).strip() for line in text.replace("\r", "\n").split("\n")]
        return "\n".join(line for line in lines if line)[:max_len]

    def _extract_receipt_date(raw_text: str) -> str | None:
        text = raw_text or ""

        # Format: DD/MM/YYYY, DD-MM-YYYY, DD.MM.YYYY
        matches = re.findall(r"(\d{1,2})[/.-](\d{1,2})[/.-](\d{2,4})", text)
        for day, month, year in matches:
            year = year if len(year) == 4 else f"20{year}"
            try:
                parsed = datetime(int(year), int(month), int(day))
                # Reject obviously wrong dates (future by more than 1 day, or before 2000)
                now = datetime.now()
                if 2000 <= parsed.year <= now.year + 1:
                    return parsed.strftime("%Y-%m-%d")
            except ValueError:
                continue

        # Format: "29 ABR 2026", "29 ABRIL 2026"
        month_pt = {
            "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
            "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
            "janeiro": 1, "fevereiro": 2, "março": 3, "abril": 4, "maio": 5,
            "junho": 6, "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10,
            "novembro": 11, "dezembro": 12,
        }
        for m in re.finditer(r"(\d{1,2})\s+([a-záéíóúãêç]+)\.?\s+(\d{4})", text, re.IGNORECASE):
            day_s, mon_s, year_s = m.group(1), m.group(2).lower(), m.group(3)
            if mon_s in month_pt:
                try:
                    parsed = datetime(int(year_s), month_pt[mon_s], int(day_s))
                    return parsed.strftime("%Y-%m-%d")
                except ValueError:
                    continue
        return None

    def _extract_receipt_amount(raw_text: str) -> float | None:
        text = raw_text or ""

        def _parse_br(raw_val: str) -> float | None:
            raw_val = raw_val.strip().replace(" ", "")
            normalized = raw_val.replace(".", "").replace(",", ".") if "," in raw_val else raw_val
            try:
                v = round(float(normalized), 2)
                return v if 0.01 <= v <= 1_000_000 else None
            except ValueError:
                return None

        # Priority patterns — try from most to least specific and return first good match.
        priority_patterns = [
            r"(?:VALOR\s+PAGO|TOTAL\s+A\s+PAGAR|TOTAL\s+PAGO|VALOR\s+TOTAL)\s*:?\s*R?\$?\s*([\d.]+,[\d]{2})",
            r"(?:TOTAL|PAGAR|PAGO)\s*:?\s*R?\$?\s*([\d.]+,[\d]{2})",
            r"R\$\s*([\d.]+,[\d]{2})",
        ]
        for pattern in priority_patterns:
            for m in re.finditer(pattern, text, re.IGNORECASE):
                v = _parse_br(m.group(1))
                if v:
                    return v

        # Fallback: collect all BRL-format numbers and return the largest.
        fallback: list[float] = []
        for m in re.finditer(r"\b([\d]{1,6}[.,][\d]{2})\b", text):
            v = _parse_br(m.group(1))
            if v:
                fallback.append(v)
        return max(fallback) if fallback else None

    def _extract_receipt_merchant(raw_text: str) -> str:
        ignored_prefixes = (
            "CNPJ", "CPF", "ITEM", "QTD", "TOTAL", "SUBTOTAL", "VALOR", "CARTAO",
            "CREDITO", "DEBITO", "AUTORIZACAO", "NSU", "COD", "OPERADOR", "CAIXA",
            "VENDA", "DATA", "HORA", "PAGAMENTO", "CLIENTE", "TROCO", "BANCO",
            "DOCUMENTO", "NOTA FISCAL", "NFC-E", "NF-E", "DANFE", "SAT", "CUPOM",
            "RECIBO", "PROTOCOLO", "SERIE", "NUMERO", "VIA", "CONSULTE",
        )
        ignored_exact = (
            "CONSUMIDOR", "NAO OBRIGATORIO", "NÃO OBRIGATÓRIO", "CONSUMIDOR FINAL",
        )
        candidates: list[str] = []
        for line in (raw_text or "").splitlines():
            cleaned = sanitize_text(line.strip(), 120)
            if len(cleaned) < 5:
                continue
            upper = cleaned.upper().strip()
            if upper in ignored_exact:
                continue
            if any(upper.startswith(p) for p in ignored_prefixes):
                continue
            if re.search(r"\b\d{1,2}[/.-]\d{1,2}[/.-]\d{2,4}\b", cleaned):
                continue
            if re.fullmatch(r"[\d\s./,:-]+", cleaned):
                continue
            # Prefer lines with 2+ words (more likely a company name)
            words = [w for w in cleaned.split() if len(w) > 1]
            if len(words) >= 2:
                return cleaned
            candidates.append(cleaned)
        return candidates[0] if candidates else ""

    def _score_ocr_candidate(raw_text: str) -> int:
        score = min(len(raw_text or ""), 200)
        if _extract_receipt_amount(raw_text) is not None:
            score += 80
        if _extract_receipt_date(raw_text) is not None:
            score += 50
        if _extract_receipt_merchant(raw_text):
            score += 40
        if re.search(r"(?:TOTAL|VALOR|R\$)", raw_text or "", re.IGNORECASE):
            score += 30
        return score

    # ── OCR PT-BR normalization ────────────────────────────────────────────────

    _OCR_PTBR_FIXES: list[tuple] = [
        # Common OCR character confusions in Portuguese words
        (re.compile(r"\b0(?=\d{2,}[./])"), "O"),         # 0CNPJ → O (not actually a word but guards digits)
        (re.compile(r"(?<=[A-ZÀ-Ú])0(?=[A-ZÀ-Ú])"), "O"),  # ST0RE → STORE
        (re.compile(r"(?<=[a-zà-ú])0(?=[a-zà-ú])"), "o"),   # st0re → store
        (re.compile(r"\brn\b"), "m"),                         # rn → m
        (re.compile(r"(?<=[a-zA-ZÀ-Úà-ú])1(?=[a-zA-ZÀ-Úà-ú])"), "l"),  # va1or → valor
        (re.compile(r"\bVa1or\b", re.I), "Valor"),
        (re.compile(r"\bTota1\b", re.I), "Total"),
        (re.compile(r"\bPague\b", re.I), "Pague"),
        (re.compile(r"(?<=[A-ZÀ-Úa-zà-ú])II(?=[A-ZÀ-Úa-zà-ú])"), "ll"),  # bIIhete → bilhete
        (re.compile(r"\bCNPJ\s*/?\s*CPF\b", re.I), "CNPJ/CPF"),  # normalize spacing
    ]

    def _normalize_ocr_errors_ptbr(text: str) -> str:
        """Fix typical Tesseract character confusions in PT-BR receipts."""
        if not text:
            return text
        result = text
        for pattern, replacement in _OCR_PTBR_FIXES:
            try:
                result = pattern.sub(replacement, result)
            except Exception:
                pass
        return result

    # ── CNAE fiscal → cashflow category mapping ────────────────────────────────

    _CNAE_CATEGORY_MAP: list[tuple[str, str]] = [
        # (substring to match in cnae_fiscal_desc lowercase, category)
        ("aliment", "Alimentação"),
        ("mercearia", "Alimentação"),
        ("padaria", "Alimentação"),
        ("açougue", "Alimentação"),
        ("restaurante", "Alimentação"),
        ("lanchonete", "Alimentação"),
        ("bar e", "Alimentação"),
        ("bebida", "Alimentação"),
        ("farmácia", "Saúde"),
        ("farmacia", "Saúde"),
        ("drogaria", "Saúde"),
        ("médico", "Saúde"),
        ("médica", "Saúde"),
        ("saúde", "Saúde"),
        ("hospital", "Saúde"),
        ("clínica", "Saúde"),
        ("odontol", "Saúde"),
        ("laboratório", "Saúde"),
        ("transporte", "Transporte"),
        ("logística", "Transporte"),
        ("frete", "Transporte"),
        ("taxi", "Transporte"),
        ("combustível", "Transporte"),
        ("postos de", "Transporte"),
        ("oficina mecânica", "Transporte"),
        ("vestuário", "Vestuário"),
        ("roupa", "Vestuário"),
        ("calçado", "Vestuário"),
        ("moda", "Vestuário"),
        ("educ", "Educação"),
        ("escola", "Educação"),
        ("universidade", "Educação"),
        ("livro", "Educação"),
        ("livraria", "Educação"),
        ("serviços financeiros", "Taxas Bancárias"),
        ("banco", "Taxas Bancárias"),
        ("seguros", "Taxas Bancárias"),
        ("energia elétrica", "Contas"),
        ("gás", "Contas"),
        ("água", "Contas"),
        ("telecomunicações", "Assinaturas"),
        ("internet", "Assinaturas"),
        ("telefonia", "Assinaturas"),
        ("streaming", "Assinaturas"),
        ("aluguel", "Moradia"),
        ("condomínio", "Moradia"),
        ("construção", "Moradia"),
        ("imobiliária", "Moradia"),
        ("entretenimento", "Lazer"),
        ("cinema", "Lazer"),
        ("teatro", "Lazer"),
        ("hotel", "Lazer"),
        ("turismo", "Lazer"),
    ]

    def _cnae_to_category(cnae_desc: str) -> str | None:
        """Map a CNAE fiscal description to a cashflow category."""
        lower = (cnae_desc or "").lower()
        for keyword, cat in _CNAE_CATEGORY_MAP:
            if keyword in lower:
                return cat
        return None

    def _lookup_cnpj_data(cnpj: str) -> dict | None:
        """Fetch name + CNAE from ReceitaWS. Returns {name, category} or None."""
        try:
            import requests as _req  # noqa: PLC0415
            digits = re.sub(r"\D", "", cnpj)
            resp = _req.get(
                f"https://www.receitaws.com.br/v1/cnpj/{digits}",
                timeout=2,
                headers={"User-Agent": "Mozilla/5.0"},
            )
            if resp.status_code == 200:
                data = resp.json()
                name = str(data.get("fantasia") or data.get("nome") or "").strip()
                cnae_desc = str(data.get("cnae_fiscal_descricao") or data.get("cnae_fiscal_desc") or "").strip()
                return {
                    "name": name[:100] if name else None,
                    "category": _cnae_to_category(cnae_desc),
                    "cnae": cnae_desc[:120] if cnae_desc else None,
                }
        except Exception:
            pass
        return None

    # ── Per-field confidence ───────────────────────────────────────────────────

    def _compute_field_confidence(
        raw_text: str,
        date: str | None,
        amount: float | None,
        merchant: str | None,
        category: str | None,
    ) -> dict[str, float]:
        """
        Return a confidence score (0.0–1.0) for each extracted field.
        High = extracted cleanly, Low = guessed / missing.
        """
        fc: dict[str, float] = {}
        # Date: 1.0 if explicit DD/MM/YYYY match, 0.5 if month-name match, 0.0 if absent
        if date is None:
            fc["date"] = 0.0
        elif re.search(r"\d{2}/\d{2}/\d{4}", raw_text or ""):
            fc["date"] = 1.0
        else:
            fc["date"] = 0.6
        # Amount: 1.0 if VALOR PAGO / TOTAL label present, 0.7 if generic R$ match, 0.0 absent
        if amount is None:
            fc["amount"] = 0.0
        elif re.search(r"(?:VALOR\s+PAGO|TOTAL\s+PAGO)", raw_text or "", re.I):
            fc["amount"] = 1.0
        elif re.search(r"\bTOTAL\b", raw_text or "", re.I):
            fc["amount"] = 0.85
        else:
            fc["amount"] = 0.6
        # Merchant: 1.0 if ≥2 words, no noise prefixes, 0.5 if single word, 0.0 if absent
        if not merchant:
            fc["merchant"] = 0.0
        elif len(merchant.split()) >= 2:
            fc["merchant"] = 0.9
        else:
            fc["merchant"] = 0.5
        # Category: 1.0 if matched via CNAE, 0.7 keyword match, 0.0 if None
        if not category:
            fc["category"] = 0.0
        else:
            fc["category"] = 0.7  # will be overridden to 1.0 by caller if CNAE used
        return fc

    def _detect_receipt_type(text: str) -> str:
        """Identify the type of receipt/document from its text."""
        normalized = (text or "").lower()
        if re.search(r"nfc-?e|nota fiscal eletr[oô]nica de consumidor|sat fiscal", normalized):
            return "nfc-e"
        if re.search(r"nota fiscal eletr[oô]nica|nf-?e|danfe", normalized):
            return "nf-e"
        if re.search(r"boleto|linha digit[aá]vel|nosso n[uú]mero", normalized):
            return "boleto"
        if re.search(r"\bpix\b|chave pix|recibo pix|comprovante pix", normalized):
            return "pix"
        if re.search(r"extrato|saldo anterior|saldo atual|hist[oó]rico de transa", normalized):
            return "extrato"
        if re.search(r"cupom fiscal|ecf\b|sat\b", normalized):
            return "cupom-fiscal"
        return "recibo"

    def _extract_cnpj(text: str) -> str | None:
        """Extract first valid CNPJ from text."""
        for raw in re.findall(r"\d{2}[\.\s]?\d{3}[\.\s]?\d{3}[/\s]?\d{4}[-\s]?\d{2}", text or ""):
            digits = re.sub(r"\D", "", raw)
            if len(digits) == 14:
                return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:14]}"
        return None

    def _lookup_cnpj_name(cnpj: str) -> str | None:
        """Try to fetch company name from ReceitaWS (2 s timeout, fails silently)."""
        try:
            import requests as _req  # noqa: PLC0415
            digits = re.sub(r"\D", "", cnpj)
            resp = _req.get(
                f"https://www.receitaws.com.br/v1/cnpj/{digits}",
                timeout=2,
                headers={"User-Agent": "Mozilla/5.0"},
            )
            if resp.status_code == 200:
                data = resp.json()
                name = str(data.get("fantasia") or data.get("nome") or "").strip()
                return name[:100] if name else None
        except Exception:
            pass
        return None

    def _extract_payment_method(text: str) -> str | None:
        """Identify payment method from receipt text."""
        normalized = (text or "").lower()
        if re.search(r"\bpix\b", normalized):
            return "PIX"
        if re.search(r"mastercard|visa|elo\b|hipercard|amex|american express", normalized):
            # Check if it's credit or debit based on context
            if re.search(r"cr[eé]dito", normalized):
                return "Cartão de Crédito"
            if re.search(r"d[eé]bito", normalized):
                return "Cartão de Débito"
            return "Cartão"
        if re.search(r"cart[aã]o\s+cr[eé]dito|cr[eé]dito", normalized):
            return "Cartão de Crédito"
        if re.search(r"cart[aã]o\s+d[eé]bito|d[eé]bito", normalized):
            return "Cartão de Débito"
        if re.search(r"dinheiro|esp[eé]cie|troco", normalized):
            return "Dinheiro"
        if re.search(r"boleto", normalized):
            return "Boleto"
        if re.search(r"\bted\b|\bdoc\b|transfer[eê]ncia banc", normalized):
            return "TED/DOC"
        return None

    def _extract_receipt_items(text: str) -> list[dict]:
        """Extract product line items from receipt text (max 20)."""
        items: list[dict] = []
        # Matches: qty  DESCRIPTION  UNIT  TOTAL  e.g. "1 PAO DE ALHO 450G  13,90  13,90"
        item_re = re.compile(
            r"^(\d{1,3})\s{1,4}(.{3,40?}?)\s+([\d.]+,\d{2})\s+([\d.]+,\d{2})\s*$"
        )
        simple_re = re.compile(
            r"^(.{4,40}?)\s+([\d.]+,\d{2})\s*$"
        )
        skip_prefixes = (
            "total", "subtotal", "valor", "desconto", "troco", "pago", "taxa",
            "cnpj", "cpf", "data", "hora", "forma", "pagamento", "obrigado",
        )
        for line in (text or "").splitlines():
            stripped = line.strip()
            if len(stripped) < 5:
                continue
            if stripped.lower().startswith(skip_prefixes):
                continue
            # Try qty + desc + unit + total
            m = item_re.match(stripped)
            if m:
                try:
                    qty = int(m.group(1))
                    desc = m.group(2).strip()
                    total = float(m.group(4).replace(".", "").replace(",", "."))
                    items.append({"qty": qty, "description": desc[:60], "total": round(total, 2)})
                    if len(items) >= 20:
                        break
                    continue
                except (ValueError, ZeroDivisionError):
                    pass
            # Try simple: desc + price
            m2 = simple_re.match(stripped)
            if m2:
                desc = m2.group(1).strip()
                if re.search(r"[A-Za-záéíóúçã]", desc):  # must have letters
                    try:
                        total = float(m2.group(2).replace(".", "").replace(",", "."))
                        if 0.01 <= total <= 10000:
                            items.append({"qty": 1, "description": desc[:60], "total": round(total, 2)})
                            if len(items) >= 20:
                                break
                    except ValueError:
                        pass
        return items

    def _pdf_bytes_to_image_bytes(pdf_bytes: bytes) -> bytes | None:
        """Convert first page of a PDF to PNG bytes. Requires pdf2image + poppler."""
        try:
            from pdf2image import convert_from_bytes  # type: ignore[import-not-found]  # noqa: PLC0415
            import io as _io  # noqa: PLC0415
            images = convert_from_bytes(pdf_bytes, first_page=1, last_page=1, dpi=200)
            if images:
                buf = _io.BytesIO()
                images[0].save(buf, format="PNG")
                return buf.getvalue()
        except Exception:
            pass
        return None

    def _pdf_bytes_to_all_images(pdf_bytes: bytes) -> list[bytes]:
        """Convert every page of a PDF to PNG bytes (max 10 pages). Returns [] on failure."""
        try:
            from pdf2image import convert_from_bytes  # type: ignore[import-not-found]  # noqa: PLC0415
            import io as _io  # noqa: PLC0415
            images = convert_from_bytes(pdf_bytes, first_page=1, last_page=10, dpi=200)
            result = []
            for img in images:
                buf = _io.BytesIO()
                img.save(buf, format="PNG")
                result.append(buf.getvalue())
            return result
        except Exception:
            pass
        return []

    def _extract_text_from_receipt_image(raw_bytes: bytes) -> str:
        import io as _io
        import pytesseract  # noqa: PLC0415
        from PIL import Image, ImageFilter, ImageOps  # noqa: PLC0415

        if not raw_bytes:
            raise ValueError("Image bytes are empty")
        
        try:
            img = Image.open(_io.BytesIO(raw_bytes))
            img.load()
        except Exception as e:
            raise ValueError(f"Falha ao abrir imagem (pode estar corrompida): {e}")

        try:
            base = ImageOps.exif_transpose(img).convert("L")
        except Exception as e:
            raise ValueError(f"Falha ao processar imagem: {e}")

        # Auto-correct rotation using Tesseract OSD (best-effort, ignore failures)
        try:
            import pytesseract as _pt_rot  # noqa: PLC0415
            osd = _pt_rot.image_to_osd(base, config="--psm 0 -c min_characters_to_try=5")
            rot_m = re.search(r"Rotate:\s*(\d+)", osd)
            if rot_m:
                angle = int(rot_m.group(1))
                if angle in (90, 180, 270):
                    base = base.rotate(angle, expand=True)
        except Exception:
            pass

        # Deskew: correct continuous skew angle (not just 90° steps)
        # Uses horizontal projection method — rotate by the angle that maximises row variance.
        try:
            import numpy as _np_deskew  # noqa: PLC0415
            arr_d = _np_deskew.array(base)
            best_angle = 0.0
            best_var = -1.0
            for deg in range(-10, 11):  # search ±10°
                from PIL import Image as _PI  # noqa: PLC0415
                rotated_arr = _np_deskew.array(
                    base.rotate(deg, expand=False, fillcolor=255)
                )
                row_sums = rotated_arr.sum(axis=1).astype(float)
                var = float(row_sums.var())
                if var > best_var:
                    best_var = var
                    best_angle = float(deg)
            if abs(best_angle) >= 1.0:
                base = base.rotate(best_angle, expand=False, fillcolor=255)
        except Exception:
            pass

        # Cap resolution to max 1800px on the longest side to keep Tesseract fast.
        _MAX_DIM = 1800
        w, h = base.size
        if max(w, h) < _MAX_DIM // 2:
            # Small image: upscale 2x so characters are legible
            base = base.resize((w * 2, h * 2), Image.Resampling.LANCZOS)
        elif max(w, h) > _MAX_DIM:
            scale = _MAX_DIM / max(w, h)
            base = base.resize((int(w * scale), int(h * scale)), Image.Resampling.LANCZOS)

        # Otsu binarization: improves contrast on receipts with uneven lighting
        try:
            import numpy as _np  # noqa: PLC0415
            arr = _np.array(base)
            # Otsu threshold
            hist, bins = _np.histogram(arr.ravel(), bins=256, range=(0, 256))
            total = arr.size
            sum_total = float(_np.dot(_np.arange(256), hist))
            sum_bg = 0.0
            weight_bg = 0
            max_var = 0.0
            threshold = 128
            for i in range(256):
                weight_bg += int(hist[i])
                if weight_bg == 0:
                    continue
                weight_fg = total - weight_bg
                if weight_fg == 0:
                    break
                sum_bg += float(i * hist[i])
                mean_bg = sum_bg / weight_bg
                mean_fg = (sum_total - sum_bg) / weight_fg
                var = weight_bg * weight_fg * (mean_bg - mean_fg) ** 2
                if var > max_var:
                    max_var = var
                    threshold = i
            bw = _np.where(arr >= threshold, 255, 0).astype("uint8")
            binarized = Image.fromarray(bw)
        except Exception:
            binarized = base  # numpy unavailable or error → skip

        sharpened = ImageOps.autocontrast(binarized).filter(ImageFilter.SHARPEN)
        enlarged = base  # alias kept for the loop below

        def _is_good_enough(text: str) -> bool:
            return (
                _extract_receipt_amount(text) is not None
                and (_extract_receipt_date(text) is not None or _extract_receipt_merchant(text) != "")
            )

        # Strategy: try configs in order of Portuguese language preference.
        # OEM 2 (legacy) often works better for Portuguese than OEM 3 (neural).
        # PSM 6 = uniform blocks, PSM 11 = sparse text (good for receipts).
        configs = (
            "--oem 2 --psm 11",  # Legacy engine, sparse text (best for receipts)
            "--oem 2 --psm 6",   # Legacy engine, uniform blocks
            "--oem 3 --psm 6",   # Neural engine, uniform blocks (fallback)
        )
        candidates: list[str] = []

        # Try best Portuguese configs with best image variant first, return early if good.
        for config in configs:
            for variant in (sharpened, enlarged, base):
                try:
                    text = pytesseract.image_to_string(variant, lang="por+eng", config=config)
                except Exception:
                    continue
                normalized = _normalize_ocr_text(text)
                if normalized:
                    candidates.append(normalized)
                    if _is_good_enough(normalized):
                        return normalized

        # Word-level confidence fallback: rebuild text from only high-confidence words
        try:
            df = pytesseract.image_to_data(
                sharpened,
                lang="por+eng",
                config="--oem 2 --psm 11",
                output_type=pytesseract.Output.DATAFRAME,
            )
            mask = (df["conf"] >= 30) & df["text"].str.strip().astype(bool)
            if mask.any():
                words_df = df.loc[mask, ["line_num", "text"]].copy()
                wc_lines: list[str] = []
                for _ln, grp in words_df.groupby("line_num"):
                    wc_lines.append(" ".join(grp["text"].tolist()))
                wc_text = _normalize_ocr_text("\n".join(wc_lines))
                if wc_text:
                    candidates.append(wc_text)
        except Exception:
            pass

        return max(candidates, key=_score_ocr_candidate) if candidates else ""

    @app.post("/api/finance/cashflow/ocr")
    @limiter.limit("10/minute")
    @require_finance_key
    def finance_cashflow_ocr():
        """OCR de comprovante: imagem → campos pré-preenchidos para lançamento."""
        f = request.files.get("file")
        manual_text = _sanitize_ocr_manual_text(str(request.form.get("manual_text", "")), 2000)
        force_reanalyze = str(request.form.get("force", "")).lower() in ("1", "true", "yes")
        if not f and not manual_text:
            return jsonify({"ok": False, "error": "Envie arquivo no campo 'file' ou texto em 'manual_text'"}), 400

        raw_bytes = b""
        cache_key = ""
        is_pdf = False
        if f:
            raw_bytes = f.read(5 * 1024 * 1024)  # max 5 MB
            if len(raw_bytes) < 100 and not manual_text:
                return jsonify({"ok": False, "error": "Arquivo muito pequeno ou vazio"}), 400
            # Detect PDF by magic bytes
            if raw_bytes[:4] == b"%PDF":
                is_pdf = True
                all_pages = _pdf_bytes_to_all_images(raw_bytes)
                if all_pages:
                    # OCR every page and pick the one with the best extraction score
                    best_bytes = all_pages[0]
                    best_score = -1
                    for page_bytes in all_pages:
                        try:
                            page_text = _extract_text_from_receipt_image(page_bytes)
                            score = _score_ocr_candidate(page_text)
                            if score > best_score:
                                best_score = score
                                best_bytes = page_bytes
                        except Exception:
                            pass
                    raw_bytes = best_bytes
                    is_pdf = False
                else:
                    return jsonify({
                        "ok": False,
                        "error": "PDF recebido mas pdf2image/poppler não estão instalados no servidor. Use uma imagem JPG/PNG.",
                    }), 501
            cache_key = hashlib.sha256(raw_bytes).hexdigest()

        # Hot-path cache: repeated retries of the same receipt return immediately.
        if cache_key and not force_reanalyze:
            _cleanup_ocr_cache()
            with _ocr_cache_lock:
                cached_row = _ocr_cache.get(cache_key)
            if cached_row:
                return jsonify({
                    **cached_row.get("payload", {}),
                    "from_cache": True,
                })

        try:
            import pytesseract  # noqa: PLC0415
            from PIL import Image  # noqa: PLC0415
        except ImportError:
            if manual_text:
                normalized_manual_text = _normalize_ocr_text(manual_text)
                inferred_category = _infer_cashflow_category_from_text(manual_text)
                inferred_entry_type = _infer_cashflow_entry_type_from_text(
                    normalized_manual_text,
                    inferred_category,
                )
                lines = [l.strip() for l in normalized_manual_text.splitlines() if l.strip()]
                parsed_date = _extract_receipt_date(normalized_manual_text)
                parsed_amount = _extract_receipt_amount(normalized_manual_text)
                merchant = _extract_receipt_merchant(normalized_manual_text)
                description = sanitize_text(merchant or (lines[1][:100] if len(lines) > 1 else ""), 100)
                confidence = 0.45
                if parsed_amount is not None:
                    confidence += 0.2
                if parsed_date is not None:
                    confidence += 0.15
                if merchant:
                    confidence += 0.1
                if inferred_category:
                    confidence += 0.1
                return jsonify({
                    "ok": True,
                    "fallback_used": "manual_text",
                    "date": parsed_date,
                    "amount": parsed_amount,
                    "description": description,
                    "merchant": merchant,
                    "category": inferred_category,
                    "entry_type": inferred_entry_type,
                    "confidence": round(min(0.99, confidence), 2),
                    "raw_text": normalized_manual_text[:1000],
                })
            return jsonify({
                "ok": False,
                "pytesseract_missing": True,
                "error": "Instale pytesseract e Pillow para OCR automático: pip install pytesseract Pillow",
            }), 501

        raw_text = manual_text
        if f:
            try:
                from PIL import Image as PILImage  # noqa: PLC0415
                import io as _io_verify  # noqa: PLC0415
                
                # Validate image before OCR
                try:
                    test_img = PILImage.open(_io_verify.BytesIO(raw_bytes))
                    test_img.verify()
                except Exception as img_err:
                    if not manual_text:
                        return jsonify({
                            "ok": False,
                            "error": f"Arquivo não é uma imagem válida. Envie JPG ou PNG. Detalhes: {img_err}",
                        }), 422
                    raw_text = manual_text
                    return jsonify({
                        "ok": True,
                        "fallback_used": "manual_text",
                        "image_error": str(img_err),
                        "date": _extract_receipt_date(manual_text),
                        "amount": _extract_receipt_amount(manual_text),
                        "description": sanitize_text(
                            _extract_receipt_merchant(manual_text) or "", 100
                        ),
                        "category": _infer_cashflow_category_from_text(manual_text),
                        "entry_type": _infer_cashflow_entry_type_from_text(manual_text),
                        "confidence": 0.5,
                    })
                
                extracted_text = _extract_text_from_receipt_image(raw_bytes)
                raw_text = "\n".join(part for part in [extracted_text, manual_text] if part).strip()
            except Exception as exc:
                if not manual_text:
                    import traceback
                    return jsonify({
                        "ok": False,
                        "error": f"Erro ao processar imagem: {exc}",
                        "debug": traceback.format_exc()[:500],
                    }), 422
                raw_text = manual_text

        raw_text = _normalize_ocr_errors_ptbr(_normalize_ocr_text(raw_text))

        entry_date = _extract_receipt_date(raw_text)
        amount = _extract_receipt_amount(raw_text)

        lines = [l.strip() for l in raw_text.splitlines() if l.strip() and len(l.strip()) > 3]
        merchant = _extract_receipt_merchant(raw_text)
        description = sanitize_text(merchant or (lines[1][:100] if len(lines) > 1 else ""), 100)
        category = _infer_cashflow_category_from_text("\n".join(lines[:8]))
        entry_type = _infer_cashflow_entry_type_from_text(raw_text, category)
        receipt_type = _detect_receipt_type(raw_text)
        cnpj = _extract_cnpj(raw_text)
        payment_method = _extract_payment_method(raw_text)
        items = _extract_receipt_items(raw_text)

        # CNPJ lookup: get company name + CNAE-based category
        cnpj_data: dict | None = None
        cnae_category: str | None = None
        if cnpj:
            cnpj_data = _lookup_cnpj_data(cnpj)
            if cnpj_data:
                if not merchant and cnpj_data.get("name"):
                    merchant = cnpj_data["name"]
                    description = sanitize_text(merchant, 100)
                if cnpj_data.get("category"):
                    cnae_category = cnpj_data["category"]
                    category = cnae_category  # CNAE overrides keyword inference

        # Suggestion from previous scans with same CNPJ
        suggestion: dict | None = None
        if cnpj:
            try:
                prev = repo.get_fin_ocr_suggestion_by_cnpj(cnpj)
                if prev:
                    suggestion = {
                        "merchant": prev.get("merchant"),
                        "category": prev.get("category"),
                        "entry_type": prev.get("entry_type"),
                        "payment_method": prev.get("payment_method"),
                    }
                    # Only apply suggestion for fields not already extracted
                    if not merchant and suggestion.get("merchant"):
                        merchant = suggestion["merchant"]
                        description = sanitize_text(merchant, 100)
                    if not category and suggestion.get("category"):
                        category = suggestion["category"]
            except Exception:
                pass

        # Per-field confidence
        field_confidence = _compute_field_confidence(raw_text, entry_date, amount, merchant, category)
        if cnae_category:
            field_confidence["category"] = 1.0

        # Duplicate detection
        duplicate_warning: dict | None = None
        if amount is not None and entry_date:
            try:
                dupes = repo.find_fin_cashflow_duplicates(amount=amount, entry_date=entry_date)
                if dupes:
                    duplicate_warning = {
                        "count": len(dupes),
                        "ids": [d["id"] for d in dupes[:3]],
                        "samples": [
                            {"id": d["id"], "date": d.get("entry_date"), "desc": d.get("description")}
                            for d in dupes[:2]
                        ],
                    }
            except Exception:
                pass

        # Budget alert: check if the inferred category is near/over budget this month
        budget_alert: dict | None = None
        if category and amount is not None and entry_type == "expense" and entry_date:
            try:
                month_str = str(entry_date)[:7]
                budget = repo.get_fin_cashflow_budget(month_str)
                limit_val = budget.get(category)
                if limit_val:
                    analytics = repo.get_fin_cashflow_analytics(month=month_str)
                    expense_cats = analytics.get("categories", {}).get("expense", [])
                    spent = next(
                        (float(r["amount"]) for r in expense_cats if r["category"] == category),
                        0.0,
                    )
                    would_be = spent + float(amount)
                    pct = round(would_be / limit_val * 100 if limit_val > 0 else 0)
                    if pct >= 80:
                        budget_alert = {
                            "category": category,
                            "spent": round(spent, 2),
                            "would_be": round(would_be, 2),
                            "limit": round(limit_val, 2),
                            "pct": pct,
                            "over": would_be > limit_val,
                        }
            except Exception:
                pass

        confidence = 0.45
        if f and raw_text:
            confidence += 0.1
        if amount is not None:
            confidence += 0.2
        if entry_date is not None:
            confidence += 0.15
        if merchant:
            confidence += 0.1
        if category:
            confidence += 0.1

        payload = {
            "ok": True,
            "date": entry_date,
            "amount": amount,
            "description": description,
            "merchant": merchant,
            "category": category,
            "entry_type": entry_type,
            "confidence": round(min(0.99, confidence), 2),
            "field_confidence": field_confidence,
            "receipt_type": receipt_type,
            "cnpj": cnpj,
            "payment_method": payment_method,
            "items": items,
            "suggestion": suggestion,
            "duplicate_warning": duplicate_warning,
            "budget_alert": budget_alert,
            "fallback_used": "manual_text" if manual_text and not f else None,
            "raw_text": raw_text[:1000],
        }

        if cache_key:
            with _ocr_cache_lock:
                _ocr_cache[cache_key] = {
                    "created_ts": time.time(),
                    "payload": payload,
                }

        # Persist to DB (best-effort, non-blocking)
        try:
            repo.add_fin_ocr_scan({
                "image_hash": cache_key or None,
                "merchant": merchant,
                "cnpj": cnpj,
                "amount": amount,
                "entry_date": entry_date,
                "category": category,
                "entry_type": entry_type,
                "receipt_type": receipt_type,
                "payment_method": payment_method,
                "confidence": payload["confidence"],
                "raw_text": raw_text[:2000],
                "payload": {k: v for k, v in payload.items() if k not in ("raw_text", "items")},
            })
        except Exception:
            pass

        # Also update in-memory history (for quick access without DB query)
        with _ocr_history_lock:
            _ocr_history.insert(0, {
                "ts": time.time(),
                "merchant": merchant,
                "amount": amount,
                "date": entry_date,
                "receipt_type": receipt_type,
                "confidence": payload["confidence"],
                "cache_key": cache_key,
            })
            del _ocr_history[20:]

        return jsonify(payload)

    @app.get("/api/finance/cashflow/ocr/history")
    @limiter.limit("30/minute")
    @require_finance_key
    def finance_cashflow_ocr_history():
        """Return the last 50 OCR scans (persisted to DB)."""
        try:
            rows = repo.list_fin_ocr_scans(limit=50)
            return jsonify(rows)
        except Exception:
            # Fallback to in-memory if DB not yet migrated
            with _ocr_history_lock:
                history = list(_ocr_history)
            return jsonify([
                {
                    "ts": row["ts"],
                    "merchant": row.get("merchant"),
                    "amount": row.get("amount"),
                    "date": row.get("date"),
                    "receipt_type": row.get("receipt_type"),
                    "confidence": row.get("confidence"),
                }
                for row in history
            ])

    @app.get("/api/finance/cashflow/<int:entry_id>/attachments")
    @limiter.limit("30/minute")
    def finance_cashflow_attachments_list(entry_id: int):
        entry = repo.get_fin_cashflow_entry(entry_id)
        if not entry:
            return jsonify({"error": "Lançamento não encontrado"}), 404
        return jsonify(repo.list_fin_cashflow_attachments(entry_id))

    @app.post("/api/finance/cashflow/<int:entry_id>/attachments")
    @limiter.limit("20/minute")
    @require_finance_key
    def finance_cashflow_attachments_upload(entry_id: int):
        entry = repo.get_fin_cashflow_entry(entry_id)
        if not entry:
            return jsonify({"error": "Lançamento não encontrado"}), 404

        f = request.files.get("file")
        if not f or not f.filename:
            return jsonify({"error": "Envie o arquivo no campo 'file'"}), 400

        raw_name = sanitize_text(str(f.filename), 180).strip() or "anexo.bin"
        safe_name = re.sub(r"[^\w\-. ]", "_", raw_name)
        blob = f.read(2 * 1024 * 1024)
        if not blob:
            return jsonify({"error": "Arquivo vazio"}), 400
        if len(blob) > 2 * 1024 * 1024:
            return jsonify({"error": "Arquivo excede 2 MB"}), 400

        mime_type = sanitize_text(str(f.mimetype or "application/octet-stream"), 120).strip()
        attachment_id = repo.add_fin_cashflow_attachment(
            entry_id=entry_id,
            file_name=safe_name,
            mime_type=mime_type,
            file_blob=blob,
        )
        _audit(
            "add",
            "cashflow_attachment",
            attachment_id,
            {
                "entry_id": entry_id,
                "file_name": safe_name,
                "file_size": len(blob),
                "mime_type": mime_type,
            },
        )
        return jsonify({"ok": True, "id": attachment_id}), 201

    @app.get("/api/finance/cashflow/attachments/<int:attachment_id>/download")
    @limiter.limit("30/minute")
    def finance_cashflow_attachments_download(attachment_id: int):
        row = repo.get_fin_cashflow_attachment(attachment_id)
        if not row:
            return jsonify({"error": "Anexo não encontrado"}), 404

        file_name = sanitize_text(str(row.get("file_name") or "anexo.bin"), 180).strip() or "anexo.bin"
        mime_type = sanitize_text(str(row.get("mime_type") or "application/octet-stream"), 120).strip()
        blob = row.get("file_blob")
        if not isinstance(blob, (bytes, bytearray)):
            return jsonify({"error": "Conteúdo do anexo inválido"}), 500

        return app.response_class(
            bytes(blob),
            mimetype=mime_type,
            headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
        )

    @app.delete("/api/finance/cashflow/attachments/<int:attachment_id>")
    @limiter.limit("20/minute")
    @require_finance_key
    def finance_cashflow_attachments_delete(attachment_id: int):
        row = repo.get_fin_cashflow_attachment(attachment_id)
        if not row:
            return jsonify({"error": "Anexo não encontrado"}), 404
        repo.delete_fin_cashflow_attachment(attachment_id)
        _audit(
            "delete",
            "cashflow_attachment",
            attachment_id,
            {
                "entry_id": int(row.get("entry_id") or 0),
                "file_name": row.get("file_name"),
                "file_size": int(row.get("file_size") or 0),
            },
        )
        return jsonify({"ok": True})

    @app.get("/api/finance/cashflow/closing-pdf")
    @limiter.limit("10/minute")
    def finance_cashflow_closing_pdf():
        month = sanitize_text(str(request.args.get("month", "")), 7).strip()
        if not month:
            month = datetime.now(timezone.utc).strftime("%Y-%m")
        if not re.match(r"^\d{4}-\d{2}$", month):
            return jsonify({"error": "month inválido (use YYYY-MM)"}), 400
        theme = sanitize_text(str(request.args.get("theme", "light")), 10).strip().lower()
        if theme not in ("light", "dark"):
            theme = "light"

        summary = repo.get_fin_cashflow_summary(months=18)
        analytics = repo.get_fin_cashflow_analytics(month=month)
        budget = analytics.get("budget", {})
        totals = analytics.get("totals", {})

        month_rows = [
            row for row in (summary.get("monthly") or []) if str(row.get("month") or "") == month
        ]
        month_row = month_rows[0] if month_rows else {}

        lines = [
            "Fechamento Mensal - Fluxo de Caixa",
            f"Mes de referencia: {month}",
            f"Gerado em: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            "",
            "Resumo:",
            f"- Receitas: R$ {float(totals.get('income') or 0):,.2f}",
            f"- Despesas: R$ {float(totals.get('expense') or 0):,.2f}",
            f"- Saldo: R$ {float(totals.get('balance') or 0):,.2f}",
            f"- Taxa de poupanca: {float(totals.get('savings_rate_pct') or 0):.2f}%",
            "",
            "Orcamento:",
            f"- Total gasto: R$ {float(budget.get('total_spent') or 0):,.2f}",
            f"- Total limite: R$ {float(budget.get('total_limit') or 0):,.2f}",
            f"- Restante: R$ {float(budget.get('total_remaining') or 0):,.2f}",
            "",
            "Top categorias de gasto:",
        ]
        for row in (analytics.get("top_expenses") or [])[:5]:
            lines.append(f"- {row.get('category')}: R$ {float(row.get('amount') or 0):,.2f}")

        lines.extend([
            "",
            "Comparativo do mes (serie mensal):",
            f"- Receita mensal: R$ {float(month_row.get('income') or 0):,.2f}",
            f"- Despesa mensal: R$ {float(month_row.get('expense') or 0):,.2f}",
            f"- Saldo mensal: R$ {float(month_row.get('balance') or 0):,.2f}",
        ])

        pdf_bytes = _build_simple_pdf(lines, theme=theme)
        return app.response_class(
            pdf_bytes,
            mimetype="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=fechamento-{month}.pdf",
            },
        )

    @app.get("/api/finance/global-search")
    @limiter.limit("60/minute")
    def finance_global_search():
        """Search across cashflow entries, assets, watchlist and goals."""
        q = sanitize_text(str(request.args.get("q", "")), 120).strip()
        limit = min(int(request.args.get("limit", "20")), 50)
        search_type = sanitize_text(str(request.args.get("type", "")), 20).strip().lower() or None
        date_from = sanitize_text(str(request.args.get("date_from", "")), 10).strip() or None
        date_to = sanitize_text(str(request.args.get("date_to", "")), 10).strip() or None
        min_value_raw = sanitize_text(str(request.args.get("min_value", "")), 24).strip()
        max_value_raw = sanitize_text(str(request.args.get("max_value", "")), 24).strip()
        min_value = None
        max_value = None
        try:
            if min_value_raw:
                min_value = float(min_value_raw)
        except ValueError:
            return jsonify({"error": "min_value inválido"}), 400
        try:
            if max_value_raw:
                max_value = float(max_value_raw)
        except ValueError:
            return jsonify({"error": "max_value inválido"}), 400

        if date_from and not re.match(r"^\d{4}-\d{2}-\d{2}$", date_from):
            return jsonify({"error": "date_from inválido (use YYYY-MM-DD)"}), 400
        if date_to and not re.match(r"^\d{4}-\d{2}-\d{2}$", date_to):
            return jsonify({"error": "date_to inválido (use YYYY-MM-DD)"}), 400

        if not q or len(q) < 2:
            return jsonify({"results": []})
        cache_key = (
            f"finance:global-search:{q}:{limit}:{search_type or ''}:{date_from or ''}:"
            f"{date_to or ''}:{min_value if min_value is not None else ''}:{max_value if max_value is not None else ''}"
        )
        cached = cache.get(cache_key)
        if cached is not None:
            return jsonify(cached)
        try:
            results = repo.global_finance_search(
                q=q,
                limit=limit,
                search_type=search_type,
                date_from=date_from,
                date_to=date_to,
                min_value=min_value,
                max_value=max_value,
            )
            payload = {"results": results}
            cache.set(cache_key, payload, 45)
            return jsonify(payload)
        except Exception as exc:
            logging.getLogger(__name__).error("global search: %s", exc)
            return jsonify({"results": []}), 200

    @app.post("/api/finance/cashflow/rollover")
    @limiter.limit("10/minute")
    @require_finance_key
    def finance_cashflow_rollover():
        body = request.get_json(silent=True) or {}
        source_month = sanitize_text(str(body.get("source_month", "")), 7).strip()
        target_month = sanitize_text(str(body.get("target_month", "")), 7).strip()
        raw_type = sanitize_text(str(body.get("entry_type", "all")), 12).strip().lower()

        if not re.match(r"^\d{4}-\d{2}$", source_month):
            return jsonify({"error": "source_month inválido (use YYYY-MM)"}), 400
        if not re.match(r"^\d{4}-\d{2}$", target_month):
            return jsonify({"error": "target_month inválido (use YYYY-MM)"}), 400
        if source_month == target_month:
            return jsonify({"error": "source_month e target_month devem ser diferentes"}), 400
        if raw_type not in ("all", "income", "expense"):
            return jsonify({"error": "entry_type inválido (all|income|expense)"}), 400

        entry_type = None if raw_type == "all" else raw_type

        def _target_date(src_date: str, month_key: str) -> str:
            year = int(month_key[:4])
            month = int(month_key[5:7])
            try:
                src_day = int(str(src_date)[8:10])
            except (TypeError, ValueError):
                src_day = 1

            if month == 12:
                next_month_first = datetime(year + 1, 1, 1)
            else:
                next_month_first = datetime(year, month + 1, 1)
            last_day = (next_month_first - timedelta(days=1)).day
            safe_day = max(1, min(last_day, src_day))
            return f"{year:04d}-{month:02d}-{safe_day:02d}"

        source_rows = repo.list_fin_cashflow_entries(
            month=source_month,
            entry_type=entry_type,
            limit=5000,
        )
        target_rows = repo.list_fin_cashflow_entries(
            month=target_month,
            entry_type=entry_type,
            limit=5000,
        )

        existing_signatures: set[tuple[str, float, str, str, str]] = set()
        for row in target_rows:
            entry_date = str(row.get("entry_date") or "")[:10]
            existing_signatures.add(
                (
                    str(row.get("entry_type") or "").strip().lower(),
                    round(float(row.get("amount") or 0), 2),
                    str(row.get("category") or "").strip().lower(),
                    str(row.get("description") or "").strip().lower(),
                    entry_date,
                ),
            )

        created = 0
        skipped = 0
        created_ids: list[int] = []
        for row in source_rows:
            new_date = _target_date(str(row.get("entry_date") or ""), target_month)
            signature = (
                str(row.get("entry_type") or "").strip().lower(),
                round(float(row.get("amount") or 0), 2),
                str(row.get("category") or "").strip().lower(),
                str(row.get("description") or "").strip().lower(),
                new_date,
            )
            if signature in existing_signatures:
                skipped += 1
                continue

            new_id = repo.add_fin_cashflow_entry(
                {
                    "entry_type": signature[0],
                    "amount": signature[1],
                    "category": str(row.get("category") or "").strip(),
                    "description": str(row.get("description") or "").strip(),
                    "entry_date": new_date,
                    "notes": str(row.get("notes") or "").strip(),
                },
            )
            existing_signatures.add(signature)
            created += 1
            created_ids.append(int(new_id))

        _audit(
            "rollover",
            "cashflow",
            None,
            {
                "source_month": source_month,
                "target_month": target_month,
                "entry_type": raw_type,
                "created": created,
                "skipped": skipped,
            },
        )
        return jsonify(
            {
                "ok": True,
                "source_month": source_month,
                "target_month": target_month,
                "entry_type": raw_type,
                "created": created,
                "skipped": skipped,
                "created_ids": created_ids,
            },
        )

    @app.put("/api/finance/cashflow/recurring/<int:recurring_id>")
    @limiter.limit("20/minute")
    @require_finance_key
    def finance_cashflow_recurring_update(recurring_id: int):
        body = request.get_json(silent=True) or {}
        data: dict[str, object] = {}

        if "active" in body:
            data["active"] = bool(body.get("active"))
        if "entry_type" in body:
            entry_type = sanitize_text(str(body.get("entry_type", "")).strip().lower(), 12)
            if entry_type not in ("income", "expense"):
                return jsonify({"error": "entry_type deve ser income ou expense"}), 400
            data["entry_type"] = entry_type
        if "amount" in body:
            try:
                amount = float(body.get("amount", 0))
            except (TypeError, ValueError):
                return jsonify({"error": "amount inválido"}), 400
            if not _is_finite_number(amount) or amount <= 0:
                return jsonify({"error": "amount deve ser > 0"}), 400
            data["amount"] = amount
        if "category" in body:
            data["category"] = sanitize_text(str(body.get("category", "")), 60)
        if "subcategory" in body:
            data["subcategory"] = sanitize_text(str(body.get("subcategory", "")), 60)
        if "cost_center" in body:
            data["cost_center"] = sanitize_text(str(body.get("cost_center", "")), 60)
        if "description" in body:
            data["description"] = sanitize_text(str(body.get("description", "")), 160)
        if "notes" in body:
            data["notes"] = sanitize_text(str(body.get("notes", "")), 500)
        if "tags" in body:
            data["tags"] = _normalize_tags(body.get("tags"))
        if "frequency" in body:
            frequency = sanitize_text(str(body.get("frequency", "monthly")).strip().lower(), 20)
            if frequency not in {"monthly", "quarterly", "yearly"}:
                return jsonify({"error": "frequency inválida (monthly|quarterly|yearly)"}), 400
            data["frequency"] = frequency
        if "day_rule" in body:
            day_rule = sanitize_text(str(body.get("day_rule", "exact")).strip().lower(), 20)
            if day_rule not in {"exact", "last_day", "first_weekday", "last_weekday"}:
                return jsonify({"error": "day_rule inválido"}), 400
            data["day_rule"] = day_rule
        if "day_of_month" in body:
            dom = int(body.get("day_of_month", 1))
            data["day_of_month"] = max(1, min(31, dom))
        if "start_date" in body:
            start_date = sanitize_text(str(body.get("start_date", "")), 10).strip() or None
            if start_date and not re.match(r"^\d{4}-\d{2}-\d{2}$", start_date):
                return jsonify({"error": "start_date inválida (use YYYY-MM-DD)"}), 400
            data["start_date"] = start_date
        if "end_date" in body:
            end_date = sanitize_text(str(body.get("end_date", "")), 10).strip() or None
            if end_date and not re.match(r"^\d{4}-\d{2}-\d{2}$", end_date):
                return jsonify({"error": "end_date inválida (use YYYY-MM-DD)"}), 400
            data["end_date"] = end_date

        if not data:
            return jsonify({"error": "Nenhum campo válido para atualizar"}), 400
        if not repo.update_fin_cashflow_recurring(recurring_id, data):
            return jsonify({"error": "Recorrência não encontrada"}), 404
        _audit("update", "cashflow_recurring", recurring_id, {"fields": sorted(list(data.keys()))})
        return jsonify({"ok": True})

    @app.delete("/api/finance/cashflow/recurring/<int:recurring_id>")
    @limiter.limit("20/minute")
    @require_finance_key
    def finance_cashflow_recurring_delete(recurring_id: int):
        if not repo.delete_fin_cashflow_recurring(recurring_id):
            return jsonify({"error": "Recorrência não encontrada"}), 404
        _audit("delete", "cashflow_recurring", recurring_id, None)
        return jsonify({"ok": True})

    @app.post("/api/finance/cashflow/recurring/run")
    @limiter.limit("20/minute")
    @require_finance_key
    def finance_cashflow_recurring_run():
        body = request.get_json(silent=True) or {}
        month = sanitize_text(str(body.get("month", "")), 7).strip()
        if not month:
            month = datetime.now(timezone.utc).strftime("%Y-%m")
        if not re.match(r"^\d{4}-\d{2}$", month):
            return jsonify({"error": "month inválido (use YYYY-MM)"}), 400

        result = repo.run_fin_cashflow_recurring_for_month(month)
        _audit("run", "cashflow_recurring", None, result)
        return jsonify({"ok": True, **result})

    @app.put("/api/finance/cashflow/<int:entry_id>/status")
    @limiter.limit("20/minute")
    @require_finance_key
    def finance_update_cashflow_status(entry_id: int):
        body = request.get_json(silent=True) or {}
        raw_status = sanitize_text(str(body.get("status", "")), 12).strip().lower()
        if raw_status not in ("pending", "paid"):
            return jsonify({"error": "status inválido (pending|paid)"}), 400

        entry = repo.get_fin_cashflow_entry(entry_id)
        if not entry:
            return jsonify({"error": "Lançamento não encontrado"}), 404

        settled_at = None
        if raw_status == "paid":
            settled_at = sanitize_text(
                str(
                    body.get("settled_at")
                    or entry.get("entry_date")
                    or datetime.now(timezone.utc).strftime("%Y-%m-%d")
                ),
                10,
            )
            if not re.match(r"^\d{4}-\d{2}-\d{2}$", settled_at):
                return jsonify({"error": "settled_at inválida (use YYYY-MM-DD)"}), 400

        before_status = repo.get_fin_cashflow_status(entry_id)
        repo.set_fin_cashflow_status(entry_id, raw_status, settled_at)
        after_status = repo.get_fin_cashflow_status(entry_id)
        _audit(
            "status_update",
            "cashflow",
            entry_id,
            {
                "before_status": before_status,
                "after_status": after_status,
            },
        )
        _invalidate_cashflow_cache()
        return jsonify({
            "ok": True,
            "id": entry_id,
            "status": after_status.get("status"),
            "settled_at": after_status.get("settled_at"),
        })

    @app.post("/api/finance/cashflow/reconcile-auto")
    @limiter.limit("10/minute")
    @require_finance_key
    def finance_cashflow_auto_reconcile():
        body = request.get_json(silent=True) or {}
        month = sanitize_text(str(body.get("month", "")), 7).strip()
        if month and not re.match(r"^\d{4}-\d{2}$", month):
            return jsonify({"error": "month inválido (use YYYY-MM)"}), 400

        try:
            min_score = float(body.get("min_score", 60))
        except (TypeError, ValueError):
            min_score = 60.0
        min_score = max(40.0, min(90.0, min_score))
        apply = bool(body.get("apply", False))
        suggestions = _build_reconcile_suggestions(month or None, min_score)

        if not apply:
            _cleanup_cashflow_review_cache()
            review_id = uuid4().hex
            with _reconcile_reviews_lock:
                _reconcile_reviews[review_id] = {
                    "id": review_id,
                    "created_ts": time.time(),
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "month": month,
                    "min_score": min_score,
                    "suggestions": suggestions,
                }
            return jsonify({
                "ok": True,
                "review_mode": True,
                "review_id": review_id,
                "month": month,
                "min_score": min_score,
                "candidates": len(suggestions),
                "suggestions": suggestions[:100],
            })

        selected_ids_raw = body.get("ids", [])
        selected_ids: set[int] = set()
        if isinstance(selected_ids_raw, list):
            for item in selected_ids_raw:
                try:
                    selected_ids.add(int(item))
                except (TypeError, ValueError):
                    continue

        updated = 0
        for item in suggestions:
            item_id = int(item.get("id") or 0)
            if selected_ids and item_id not in selected_ids:
                continue
            if repo.set_fin_cashflow_status(item_id, "paid", str(item.get("entry_date") or "")):
                updated += 1
        if updated > 0:
            _invalidate_cashflow_cache()

        _audit(
            "auto_reconcile",
            "cashflow",
            None,
            {
                "month": month,
                "min_score": min_score,
                "candidates": len(suggestions),
                "updated": updated,
                "apply": apply,
            },
        )
        return jsonify({
            "ok": True,
            "month": month,
            "min_score": min_score,
            "apply": apply,
            "candidates": len(suggestions),
            "updated": updated,
            "suggestions": suggestions[:50],
        })

    @app.post("/api/finance/cashflow/reconcile-auto/confirm")
    @limiter.limit("10/minute")
    @require_finance_key
    def finance_cashflow_auto_reconcile_confirm():
        body = request.get_json(silent=True) or {}
        review_id = sanitize_text(str(body.get("review_id", "")), 64).strip()
        if not review_id:
            return jsonify({"error": "review_id obrigatório"}), 400

        _cleanup_cashflow_review_cache()
        with _reconcile_reviews_lock:
            review = _reconcile_reviews.get(review_id)
        if not review:
            return jsonify({"error": "revisão expirada ou inexistente"}), 404

        ids_raw = body.get("ids", [])
        selected_ids: set[int] = set()
        if isinstance(ids_raw, list):
            for item in ids_raw:
                try:
                    selected_ids.add(int(item))
                except (TypeError, ValueError):
                    continue

        suggestions = list(review.get("suggestions") or [])
        updated = 0
        for item in suggestions:
            entry_id = int(item.get("id") or 0)
            if selected_ids and entry_id not in selected_ids:
                continue
            if repo.set_fin_cashflow_status(entry_id, "paid", str(item.get("entry_date") or "")):
                updated += 1

        if updated > 0:
            _invalidate_cashflow_cache()
        with _reconcile_reviews_lock:
            _reconcile_reviews.pop(review_id, None)

        _audit(
            "auto_reconcile_confirm",
            "cashflow",
            None,
            {
                "review_id": review_id,
                "candidates": len(suggestions),
                "selected": len(selected_ids) if selected_ids else len(suggestions),
                "updated": updated,
            },
        )
        return jsonify({
            "ok": True,
            "review_id": review_id,
            "candidates": len(suggestions),
            "updated": updated,
        })

    # ── Excel / CSV Import ──────────────────────────────────

    # Column name aliases (Portuguese / English / broker formats)
    _COL_MAP: dict[str, str] = {}
    for _alias, _field in {
        "simbolo": "symbol", "símbolo": "symbol", "symbol": "symbol",
        "ticker": "symbol", "ativo": "symbol", "codigo": "symbol",
        "código": "symbol", "papel": "symbol",
        "nome": "name", "name": "name",
        "tipo": "asset_type", "type": "asset_type", "asset_type": "asset_type",
        "tipo_ativo": "asset_type", "classe": "asset_type",
        "operacao": "tx_type", "operação": "tx_type", "tx_type": "tx_type",
        "tipo_operacao": "tx_type", "tipo_operação": "tx_type",
        "cv": "tx_type", "c/v": "tx_type",
        "quantidade": "quantity", "qtd": "quantity", "qtde": "quantity",
        "qty": "quantity", "quantity": "quantity",
        "preco": "price", "preço": "price", "price": "price",
        "valor_unitario": "price", "preco_unitario": "price",
        "preço_unitário": "price",
        "taxas": "fees", "taxa": "fees", "fees": "fees",
        "corretagem": "fees", "emolumentos": "fees",
        "data": "date", "date": "date", "data_operacao": "date",
        "data_operação": "date",
        "notas": "notes", "notes": "notes", "observacao": "notes",
        "observação": "notes", "obs": "notes",
        "moeda": "currency", "currency": "currency",
    }.items():
        _COL_MAP[_alias] = _field

    def _normalize_col(name: str) -> str:
        """Normalize a column header to a known field name."""
        clean = (
            name.strip()
            .lower()
            .replace(" ", "_")
            .replace("-", "_")
        )
        # Remove accents with simple mapping
        for a, b in [("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"),
                      ("ú", "u"), ("ã", "a"), ("õ", "o"), ("ç", "c")]:
            clean = clean.replace(a, b)
        return _COL_MAP.get(clean, clean)

    def _parse_tx_type(val: str) -> str:
        v = val.strip().lower()
        if v in ("c", "compra", "buy", "b"):
            return "buy"
        if v in ("v", "venda", "sell", "s"):
            return "sell"
        return "buy"

    def _parse_asset_type(val: str) -> str:
        v = val.strip().lower()
        type_map = {
            "acao": "stock", "ação": "stock", "stock": "stock",
            "acoes": "stock", "ações": "stock",
            "fii": "fii", "crypto": "crypto", "cripto": "crypto",
            "criptomoeda": "crypto", "bitcoin": "crypto",
            "etf": "etf", "fundo": "fund", "fund": "fund",
            "renda fixa": "renda-fixa", "renda-fixa": "renda-fixa",
            "rf": "renda-fixa", "tesouro": "renda-fixa",
        }
        return type_map.get(v, "stock")

    def _parse_number(val: str) -> float:
        """Parse number from various formats (1.234,56 or 1,234.56)."""
        v = str(val).strip().replace("R$", "").replace("$", "").strip()
        if not v or v == "-":
            return 0.0
        # Brazilian format: 1.234,56
        if "," in v and "." in v:
            if v.rindex(",") > v.rindex("."):
                v = v.replace(".", "").replace(",", ".")
            else:
                v = v.replace(",", "")
        elif "," in v:
            v = v.replace(",", ".")
        return float(v)

    @app.post("/api/finance/import")
    @limiter.limit("5/minute")
    @require_finance_key
    def finance_import():
        """Import assets & transactions from CSV or Excel file."""
        f = request.files.get("file")
        if not f or not f.filename:
            return jsonify({"error": "Envie um arquivo CSV ou Excel"}), 400

        fname = f.filename.lower()
        rows: list[dict] = []

        try:
            if fname.endswith(".csv") or fname.endswith(".tsv"):
                # CSV / TSV
                raw = f.read()
                # Try UTF-8 first, then latin-1
                try:
                    text = raw.decode("utf-8-sig")
                except UnicodeDecodeError:
                    text = raw.decode("latin-1")
                sep = "\t" if fname.endswith(".tsv") else None
                if sep is None:
                    # Auto-detect separator
                    first_line = text.split("\n")[0]
                    if "\t" in first_line:
                        sep = "\t"
                    elif ";" in first_line:
                        sep = ";"
                    else:
                        sep = ","
                reader = csv.DictReader(
                    io.StringIO(text), delimiter=sep,
                )
                for row in reader:
                    mapped = {}
                    for k, v in row.items():
                        if k and v:
                            mapped[_normalize_col(k)] = v.strip()
                    if mapped.get("symbol"):
                        rows.append(mapped)

            elif fname.endswith((".xlsx", ".xls")):
                import openpyxl

                wb = openpyxl.load_workbook(
                    io.BytesIO(f.read()), read_only=True,
                )
                ws = wb.active
                if ws is None:
                    return jsonify({"error": "Planilha sem aba ativa"}), 400
                headers: list[str] = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        headers = [
                            _normalize_col(str(c or "")) for c in row
                        ]
                        continue
                    mapped = {}
                    for j, val in enumerate(row):
                        if j < len(headers) and val is not None:
                            mapped[headers[j]] = str(val).strip()
                    if mapped.get("symbol"):
                        rows.append(mapped)
                wb.close()
            else:
                return jsonify({
                    "error": "Formato não suportado. Use .csv, .tsv ou .xlsx",
                }), 400
        except Exception as exc:
            logger.warning("Import parse error: %s", exc)
            return jsonify({"error": f"Erro ao ler arquivo: {exc}"}), 400

        if not rows:
            return jsonify({
                "error": "Nenhuma linha válida encontrada. "
                "O arquivo precisa ter ao menos uma coluna 'símbolo' ou 'symbol'.",
            }), 400

        imported = 0
        errors_list: list[str] = []
        has_tx_cols = any(
            r.get("quantity") or r.get("price") for r in rows
        )

        for idx, row in enumerate(rows, start=2):
            try:
                symbol = sanitize_text(
                    row["symbol"].upper().strip(), 20,
                )
                name = sanitize_text(
                    row.get("name", symbol), 100,
                )
                asset_type = _parse_asset_type(
                    row.get("asset_type", "stock"),
                )
                currency = sanitize_text(
                    row.get("currency", "BRL").upper(), 10,
                )

                # Upsert the asset
                asset_id = repo.upsert_fin_asset({
                    "symbol": symbol,
                    "name": name,
                    "asset_type": asset_type,
                    "currency": currency,
                })

                # If has transaction columns, also create transaction
                if has_tx_cols and row.get("quantity") and row.get("price"):
                    qty = _parse_number(row["quantity"])
                    price = _parse_number(row["price"])
                    fees = _parse_number(row.get("fees", "0"))
                    tx_type = _parse_tx_type(row.get("tx_type", "buy"))
                    tx_date = row.get("date", "")
                    if not tx_date:
                        tx_date = datetime.now().strftime("%Y-%m-%d")

                    repo.add_fin_transaction({
                        "asset_id": asset_id,
                        "tx_type": tx_type,
                        "quantity": qty,
                        "price": price,
                        "total": qty * price + fees,
                        "fees": fees,
                        "notes": sanitize_text(
                            row.get("notes", "Importado"), 500,
                        ),
                        "tx_date": tx_date,
                    })
                    _recalc_portfolio(repo, asset_id)

                imported += 1
            except Exception as exc:
                errors_list.append(f"Linha {idx}: {exc}")

        _invalidate_financial_state_cache(include_market=True)
        return jsonify({
            "ok": True,
            "imported": imported,
            "total_rows": len(rows),
            "errors": errors_list[:20],
        }), 200 if imported else 400

    # ── Import Template (sample CSV) ───────────────────────

    @app.get("/api/finance/import-template")
    def finance_import_template():
        header = "simbolo;nome;tipo;operacao;quantidade;preco;taxas;data;notas\n"
        sample = "PETR4;Petrobras PN;acao;compra;100;35.50;4.90;2026-01-15;Minha primeira compra\n"
        content = header + sample
        return (
            content,
            200,
            {
                "Content-Type": "text/csv; charset=utf-8",
                "Content-Disposition": "attachment; filename=modelo_importacao.csv",
            },
        )

    # ── Import Nota de Corretagem (brokerage note) ─────────

    @app.post("/api/finance/import-nota")
    @limiter.limit("5/minute")
    @require_finance_key
    def finance_import_nota():
        """Import brokerage note (nota de corretagem) CSV/Excel.

        Supports B3/CEI export and common broker formats.
        Expected columns (flexible mapping):
          - papel/titulo/ativo/symbol
          - tipo negociacao/operacao (C/V or Compra/Venda)
          - quantidade/qtd
          - preco/valor unitario
          - valor total/total
          - taxa corretagem/corretagem/emolumentos
          - data pregao/data/date
        """
        f = request.files.get("file")
        if not f or not f.filename:
            return jsonify({
                "error": "Envie o arquivo da nota de corretagem",
            }), 400

        fname = f.filename.lower()

        # Extra column mappings for brokerage notes
        nota_col_map = {
            "papel": "symbol", "titulo": "symbol", "ativo": "symbol",
            "cod_negociacao": "symbol", "codigo": "symbol",
            "especificacao_titulo": "name",
            "tipo_negociacao": "tx_type", "tipo_operacao": "tx_type",
            "c_v": "tx_type", "cv": "tx_type", "operacao": "tx_type",
            "quantidade": "quantity", "qtd": "quantity",
            "qtde": "quantity", "qtd_negociada": "quantity",
            "preco": "price", "preco_unitario": "price",
            "valor_unitario": "price", "pu": "price",
            "preco_ajuste": "price",
            "valor_total": "total", "valor_operacao": "total",
            "valor_liquido": "total",
            "corretagem": "fees", "taxa_corretagem": "fees",
            "emolumentos": "fees2", "taxa_liquidacao": "fees3",
            "impostos": "fees4",
            "data_pregao": "date", "data_negocio": "date",
            "data": "date", "dt_negocio": "date",
            "prazo": "notes",
            "mercado": "market",
        }

        def normalize_nota_col(col: str) -> str:
            """Normalise brokerage note column name."""
            c = col.strip().lower()
            for a, b in [("á", "a"), ("é", "e"), ("í", "i"),
                         ("ó", "o"), ("ú", "u"), ("ã", "a"),
                         ("õ", "o"), ("ç", "c")]:
                c = c.replace(a, b)
            c = c.replace(" ", "_").replace("-", "_").replace("/", "_")
            return nota_col_map.get(c, _normalize_col(c))

        rows: list[dict] = []
        try:
            if fname.endswith((".csv", ".tsv")):
                raw = f.read()
                try:
                    text = raw.decode("utf-8-sig")
                except UnicodeDecodeError:
                    text = raw.decode("latin-1")
                first_line = text.split("\n")[0]
                sep = "\t" if "\t" in first_line else (
                    ";" if ";" in first_line else ","
                )
                reader = csv.DictReader(
                    io.StringIO(text), delimiter=sep,
                )
                for row in reader:
                    mapped = {}
                    for k, v in row.items():
                        if k and v:
                            mapped[normalize_nota_col(k)] = v.strip()
                    if mapped.get("symbol"):
                        rows.append(mapped)
            elif fname.endswith((".xlsx", ".xls")):
                import openpyxl

                wb = openpyxl.load_workbook(
                    io.BytesIO(f.read()), read_only=True,
                )
                ws = wb.active
                if ws is None:
                    return jsonify({"error": "Planilha sem aba ativa"}), 400
                headers: list[str] = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        headers = [
                            normalize_nota_col(str(c or ""))
                            for c in row
                        ]
                        continue
                    mapped = {}
                    for j, val in enumerate(row):
                        if j < len(headers) and val is not None:
                            mapped[headers[j]] = str(val).strip()
                    if mapped.get("symbol"):
                        rows.append(mapped)
                wb.close()
            else:
                return jsonify({
                    "error": "Use arquivo .csv, .tsv ou .xlsx",
                }), 400
        except Exception as exc:
            logger.warning("Nota import parse: %s", exc)
            return jsonify({"error": f"Erro ao ler: {exc}"}), 400

        if not rows:
            return jsonify({
                "error": "Nenhuma operação encontrada. "
                "Verifique se o arquivo contém colunas como "
                "'papel', 'quantidade', 'preço'.",
            }), 400

        imported = 0
        errors_list: list[str] = []

        for idx, row in enumerate(rows, 2):
            try:
                symbol = sanitize_text(
                    row["symbol"].upper().strip(), 20,
                )
                # Guess asset type from symbol
                at = "stock"
                if symbol.endswith("11") and len(symbol) >= 5:
                    at = "fii"
                elif symbol.endswith("11B"):
                    at = "fii"

                name = sanitize_text(
                    row.get("name", symbol), 100,
                )
                asset_id = repo.upsert_fin_asset({
                    "symbol": symbol,
                    "name": name,
                    "asset_type": at,
                    "currency": "BRL",
                })

                qty = _parse_number(row.get("quantity", "0"))
                price = _parse_number(row.get("price", "0"))
                if qty <= 0 or price <= 0:
                    errors_list.append(
                        f"Linha {idx}: quantidade ou preço inválido",
                    )
                    continue

                # Aggregate fees from multiple fee columns
                fees = sum(
                    _parse_number(row.get(k, "0"))
                    for k in ("fees", "fees2", "fees3", "fees4")
                )

                total = _parse_number(row.get("total", "0"))
                if total <= 0:
                    total = qty * price + fees

                tx_type = _parse_tx_type(row.get("tx_type", "buy"))
                tx_date = row.get("date", "")
                if not tx_date:
                    tx_date = datetime.now().strftime("%Y-%m-%d")

                repo.add_fin_transaction({
                    "asset_id": asset_id,
                    "tx_type": tx_type,
                    "quantity": qty,
                    "price": price,
                    "total": total,
                    "fees": fees,
                    "notes": sanitize_text(
                        row.get("notes", "Importado de nota"),
                        500,
                    ),
                    "tx_date": tx_date,
                })
                _recalc_portfolio(repo, asset_id)
                imported += 1
            except Exception as exc:
                errors_list.append(f"Linha {idx}: {exc}")

        _invalidate_financial_state_cache(include_market=True)
        return jsonify({
            "ok": True,
            "imported": imported,
            "total_rows": len(rows),
            "errors": errors_list[:20],
        }), 200 if imported else 400

    # ── Import B3 Movimentação (CEI) ──────────────────────

    @app.post("/api/finance/import-b3")
    @limiter.limit("5/minute")
    @require_finance_key
    def finance_import_b3():
        """Import B3/CEI 'Movimentação' xlsx export.

        Expected columns:
          Entrada/Saída | Data | Movimentação | Produto |
          Instituição | Quantidade | Preço unitário | Valor da Operação
        """
        f = request.files.get("file")
        if not f or not f.filename:
            return jsonify({"error": "Envie o arquivo xlsx"}), 400

        fname = f.filename.lower()
        if not fname.endswith((".xlsx", ".xls")):
            return jsonify({
                "error": "Formato não suportado. Use .xlsx",
            }), 400

        # Movement types that generate BUY/SELL transactions
        TX_BUY_TYPES = {
            "transferência - liquidação",
            "bonificação em ativos",
            "fração em ativos",
        }
        TX_TRADE = "compra / venda"
        TX_SELL_TYPES = {"leilão de fração"}
        # Movement types that generate dividends
        DIV_TYPE_MAP = {
            "rendimento": "rendimento",
            "dividendo": "dividendo",
            "juros sobre capital próprio": "jcp",
            "juros sobre capital próprio - transferido": "jcp",
        }
        # Movement types to skip
        SKIP_TYPES = {
            "atualização",
            "transferência",
            "cessão de direitos",
            "cessão de direitos - solicitada",
            "direito de subscrição",
            "direitos de subscrição - não exercido",
        }

        try:
            import openpyxl

            wb = openpyxl.load_workbook(
                io.BytesIO(f.read()), read_only=False,
            )
            # Find the Movimentação sheet or use active
            ws = None
            for name in wb.sheetnames:
                if "movimenta" in name.lower():
                    ws = wb[name]
                    break
            if ws is None:
                ws = wb.active
            if ws is None:
                return jsonify({"error": "Planilha sem aba ativa"}), 400

            raw_rows: list[dict] = []
            headers: list[str] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [
                        str(c or "").strip() for c in row
                    ]
                    continue
                d: dict = {}
                for j, val in enumerate(row):
                    if j < len(headers):
                        d[headers[j]] = val
                raw_rows.append(d)
            wb.close()
        except Exception as exc:
            logger.warning("B3 import parse error: %s", exc)
            return jsonify({"error": f"Erro ao ler: {exc}"}), 400

        if not raw_rows:
            return jsonify({
                "error": "Arquivo vazio ou sem dados.",
            }), 400

        # Validate expected columns
        expected = {"Produto", "Movimentação", "Quantidade"}
        found = set(headers)
        if not expected.issubset(found):
            missing = expected - found
            return jsonify({
                "error": (
                    "Colunas esperadas não encontradas: "
                    f"{', '.join(missing)}. "
                    "Este arquivo é um export de Movimentação "
                    "do B3/CEI?"
                ),
            }), 400

        tx_imported = 0
        div_imported = 0
        skipped = 0
        errors_list: list[str] = []

        for idx, row in enumerate(raw_rows, start=2):
            try:
                mov = str(row.get("Movimentação", "")).strip()
                mov_lower = mov.lower()
                entrada = str(
                    row.get("Entrada/Saída", ""),
                ).strip().lower()
                produto = str(row.get("Produto", "")).strip()
                qty_raw = row.get("Quantidade")
                price_raw = row.get("Preço unitário")
                val_raw = row.get("Valor da Operação")

                # Skip known non-actionable types
                if mov_lower in SKIP_TYPES:
                    skipped += 1
                    continue

                # Extract ticker and name from Produto
                if " - " in produto:
                    parts = produto.split(" - ", 1)
                    symbol = parts[0].strip().upper()
                    name = parts[1].strip()
                else:
                    symbol = produto.strip().upper()
                    name = symbol

                if not symbol:
                    skipped += 1
                    continue

                # Skip subscription rights tickers (e.g. XPML12)
                # that are not actual tradeable assets
                if symbol.endswith("12") and mov_lower in (
                    "cessão de direitos",
                    "direito de subscrição",
                ):
                    skipped += 1
                    continue

                # Parse date DD/MM/YYYY → YYYY-MM-DD
                raw_date = str(row.get("Data", "")).strip()
                tx_date = ""
                if "/" in raw_date:
                    dp = raw_date.split("/")
                    if len(dp) == 3:
                        tx_date = (
                            f"{dp[2]}-{dp[1].zfill(2)}"
                            f"-{dp[0].zfill(2)}"
                        )
                if not tx_date:
                    tx_date = datetime.now().strftime("%Y-%m-%d")

                # Parse qty, price, total
                qty = _parse_number(str(qty_raw or "0"))
                price = _parse_number(str(price_raw or "0"))
                total = _parse_number(str(val_raw or "0"))

                # Auto-detect asset type
                asset_type = "stock"
                sym_upper = symbol.upper()
                if sym_upper.endswith(("11", "11B")):
                    if len(sym_upper) >= 5:
                        asset_type = "fii"
                elif sym_upper.startswith("CDB"):
                    asset_type = "renda-fixa"

                # Truncate name to 100 chars
                name = sanitize_text(name, 100)
                symbol = sanitize_text(symbol, 20)

                # ── Handle DIVIDENDS ──
                if mov_lower in DIV_TYPE_MAP:
                    if qty <= 0 or price <= 0:
                        skipped += 1
                        continue

                    div_type = DIV_TYPE_MAP[mov_lower]
                    asset_id = repo.upsert_fin_asset({
                        "symbol": symbol,
                        "name": name,
                        "asset_type": asset_type,
                        "currency": "BRL",
                    })
                    repo.add_fin_dividend({
                        "asset_id": asset_id,
                        "div_type": div_type,
                        "amount_per_share": price,
                        "total_amount": total if total > 0
                        else qty * price,
                        "quantity": qty,
                        "ex_date": tx_date,
                        "pay_date": tx_date,
                        "notes": f"Importado B3: {mov}",
                    })
                    div_imported += 1
                    continue

                # ── Handle TRANSACTIONS ──
                tx_type = None

                if mov_lower == TX_TRADE:
                    # COMPRA / VENDA: Debito=buy, Credito=sell
                    tx_type = (
                        "buy" if entrada == "debito"
                        else "sell"
                    )
                elif mov_lower in TX_BUY_TYPES:
                    tx_type = "buy"
                elif mov_lower in TX_SELL_TYPES:
                    tx_type = "sell"
                else:
                    # Unknown type, skip
                    skipped += 1
                    continue

                if qty <= 0 or price <= 0:
                    # Bonificação / fração may have price=0
                    if mov_lower in (
                        "bonificação em ativos",
                        "fração em ativos",
                    ) and qty > 0:
                        price = 0.0
                        total = 0.0
                    else:
                        skipped += 1
                        continue

                if total <= 0:
                    total = qty * price

                asset_id = repo.upsert_fin_asset({
                    "symbol": symbol,
                    "name": name,
                    "asset_type": asset_type,
                    "currency": "BRL",
                })
                repo.add_fin_transaction({
                    "asset_id": asset_id,
                    "tx_type": tx_type,
                    "quantity": qty,
                    "price": price,
                    "total": total,
                    "fees": 0,
                    "notes": f"Importado B3: {mov}",
                    "tx_date": tx_date,
                })
                _recalc_portfolio(repo, asset_id)
                tx_imported += 1

            except Exception as exc:
                errors_list.append(f"Linha {idx}: {exc}")

        _invalidate_financial_state_cache(
            include_market=True,
            include_dividends=True,
        )
        total_imported = tx_imported + div_imported
        return jsonify({
            "ok": total_imported > 0,
            "transactions": tx_imported,
            "dividends": div_imported,
            "skipped": skipped,
            "total_rows": len(raw_rows),
            "errors": errors_list[:20],
        }), 200 if total_imported else 400

    # ── Asset Price History ─────────────────────────────────

    @app.get("/api/finance/asset-history/<int:asset_id>")
    @limiter.limit("30/minute")
    def finance_asset_history_alt(asset_id: int):
        limit = min(3650, max(1, int(request.args.get("limit", "90"))))
        tx_sig = repo.get_fin_transaction_signature(asset_id)
        cache_key = (
            f"finance:asset-history:v4:{asset_id}:{limit}:"
            f"{tx_sig['tx_count']}:{tx_sig['max_id']}:{tx_sig['max_tx_date']}"
        )
        cached = cache.get(cache_key)
        if cached:
            return jsonify(cached)
        payload = repo.get_fin_asset_history(asset_id, limit)
        cache.set(cache_key, payload, FINANCE_CACHE_TTLS["asset_history"])
        return jsonify(payload)

    @app.get("/api/finance/audit")
    @limiter.limit("20/minute")
    @require_finance_key
    def finance_audit_logs():
        limit = min(300, max(1, int(request.args.get("limit", "100"))))
        target_type = sanitize_text(
            str(request.args.get("target_type", "")),
            40,
        ).strip().lower()
        action = sanitize_text(str(request.args.get("action", "")), 40).strip().lower()
        target_id = request.args.get("target_id", type=int)

        cache_key = f"finance:audit:{limit}:{target_type}:{action}:{target_id or ''}"
        cached = cache.get(cache_key)
        if cached:
            return jsonify(cached)

        payload = repo.list_fin_audit_logs(limit)
        if target_type:
            payload = [row for row in payload if str(row.get("target_type") or "").lower() == target_type]
        if action:
            payload = [row for row in payload if str(row.get("action") or "").lower() == action]
        if target_id is not None:
            payload = [row for row in payload if int(row.get("target_id") or 0) == int(target_id)]

        cache.set(cache_key, payload, FINANCE_CACHE_TTLS["audit"])
        return jsonify(payload)

    # ── Export Data ─────────────────────────────────────────

    @app.get("/api/finance/export")
    @limiter.limit("5/minute")
    def finance_export():
        """Export portfolio + transactions as CSV or Excel."""
        fmt = request.args.get("format", "csv")
        export_type = request.args.get("type", "all")

        portfolio = repo.get_fin_portfolio()
        transactions = repo.list_fin_transactions(limit=5000)
        dividends = repo.list_fin_dividends()

        if fmt == "xlsx":
            import openpyxl

            wb = openpyxl.Workbook()

            # Portfolio sheet
            if export_type in ("all", "portfolio"):
                ws = wb.active
                if ws is None:
                    return jsonify({"error": "Falha ao criar planilha"}), 500
                ws.title = "Portfólio"
                ws.append([
                    "Símbolo", "Nome", "Tipo", "Quantidade",
                    "Preço Médio", "Preço Atual", "Total Investido",
                    "Valor Atual", "P&L", "P&L %",
                ])
                for p in portfolio:
                    current_val = (p.get("current_price") or 0) * p.get("quantity", 0)
                    pnl = current_val - p.get("total_invested", 0)
                    pnl_pct = (
                        (pnl / p["total_invested"] * 100)
                        if p["total_invested"] else 0
                    )
                    ws.append([
                        p["symbol"], p.get("name", ""), p.get("asset_type", ""),
                        p["quantity"], p["avg_price"],
                        p.get("current_price", 0), p["total_invested"],
                        round(current_val, 2), round(pnl, 2),
                        round(pnl_pct, 2),
                    ])

            # Transactions sheet
            if export_type in ("all", "transactions"):
                ws2 = wb.create_sheet("Transações")
                ws2.append([
                    "Data", "Tipo", "Símbolo", "Quantidade",
                    "Preço", "Total", "Taxas", "Notas",
                ])
                for t in transactions:
                    ws2.append([
                        (t.get("tx_date") or "")[:10],
                        t["tx_type"].upper(), t["symbol"],
                        t["quantity"], t["price"], t["total"],
                        t.get("fees", 0), t.get("notes", ""),
                    ])

            # Dividends sheet
            if export_type in ("all", "dividends") and dividends:
                ws3 = wb.create_sheet("Dividendos")
                ws3.append([
                    "Data", "Símbolo", "Tipo", "Valor/Ação",
                    "Quantidade", "Total", "Notas",
                ])
                for d in dividends:
                    ws3.append([
                        (d.get("pay_date") or "")[:10],
                        d.get("symbol", ""), d.get("div_type", ""),
                        d.get("amount_per_share", 0),
                        d.get("quantity", 0),
                        d.get("total_amount", 0),
                        d.get("notes", ""),
                    ])

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return (
                buf.getvalue(),
                200,
                {
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "Content-Disposition": "attachment; filename=financeiro.xlsx",
                },
            )

        # CSV default
        output = io.StringIO()
        writer = csv.writer(output, delimiter=";")
        writer.writerow([
            "Tipo", "Data", "Símbolo", "Nome", "Classe",
            "Quantidade", "Preço", "Total", "Taxas", "Notas",
        ])
        for t in transactions:
            writer.writerow([
                t["tx_type"].upper(),
                (t.get("tx_date") or "")[:10],
                t["symbol"], t.get("name", ""), t.get("asset_type", ""),
                t["quantity"], t["price"], t["total"],
                t.get("fees", 0), t.get("notes", ""),
            ])
        for d in dividends:
            writer.writerow([
                d.get("div_type", "DIVIDEND").upper(),
                (d.get("pay_date") or "")[:10],
                d.get("symbol", ""), d.get("asset_name", ""), "",
                d.get("quantity", 0), d.get("amount_per_share", 0),
                d.get("total_amount", 0), 0, d.get("notes", ""),
            ])
        return (
            output.getvalue(),
            200,
            {
                "Content-Type": "text/csv; charset=utf-8",
                "Content-Disposition": "attachment; filename=financeiro.csv",
            },
        )

    # ── IR Report ───────────────────────────────────────────

    @app.get("/api/finance/ir-report")
    @limiter.limit("10/minute")
    def finance_ir_report():
        year = request.args.get("year", datetime.now().year, type=int)
        report = repo.get_fin_ir_report(year)
        return jsonify(report)

    # ── Market Data (brapi.dev for BR stocks, CoinGecko for crypto) ──

    @app.get("/api/finance/market-data")
    @limiter.limit("10/minute")
    def finance_market_data():
        """Fetch live quotes for tracked assets."""
        cached = cache.get("finance:market")
        if cached:
            return jsonify(cached)

        now_iso = datetime.now(timezone.utc).isoformat()

        def _resolve_brapi_monthly_limit() -> int:
            cfg_limit = app.config.get("BRAPI_MONTHLY_LIMIT")
            if cfg_limit is None:
                cfg_limit = repo.get_setting("brapi_monthly_limit", "15000")
            try:
                parsed = int(str(cfg_limit).strip())
            except Exception:
                parsed = 15000
            return max(100, min(500000, parsed))

        def _resolve_brapi_reserve_pct() -> int:
            raw = app.config.get("BRAPI_RESERVE_PCT")
            if raw is None:
                raw = repo.get_setting("brapi_reserve_pct", "15")
            try:
                parsed = int(str(raw).strip())
            except Exception:
                parsed = 15
            return max(0, min(50, parsed))

        def _resolve_brapi_max_calls_per_request() -> int:
            raw = app.config.get("BRAPI_MAX_CALLS_PER_REQUEST")
            if raw is None:
                raw = repo.get_setting("brapi_max_calls_per_request", "2")
            try:
                parsed = int(str(raw).strip())
            except Exception:
                parsed = 2
            return max(0, min(200, parsed))

        brapi_month_key = datetime.now(timezone.utc).strftime("%Y%m")
        brapi_usage_key = f"brapi_usage:{brapi_month_key}"
        try:
            brapi_usage_start = int(repo.get_setting(brapi_usage_key, "0") or 0)
        except Exception:
            brapi_usage_start = 0
        brapi_limit = _resolve_brapi_monthly_limit()
        brapi_reserve_pct = _resolve_brapi_reserve_pct()
        brapi_reserve_calls = int(math.ceil(brapi_limit * (brapi_reserve_pct / 100.0)))
        brapi_usable_limit = max(0, brapi_limit - brapi_reserve_calls)
        brapi_max_calls_per_request = _resolve_brapi_max_calls_per_request()
        brapi_usage_delta = 0
        brapi_lock = Lock()

        def _try_reserve_brapi_call() -> bool:
            nonlocal brapi_usage_delta
            with brapi_lock:
                if brapi_usage_delta >= brapi_max_calls_per_request:
                    return False
                if brapi_usage_start + brapi_usage_delta >= brapi_usable_limit:
                    return False
                brapi_usage_delta += 1
                return True

        assets = repo.list_fin_assets()
        watchlist = repo.list_fin_watchlist()

        # Collect all symbols by type
        stock_symbols = set()
        crypto_ids = set()
        for item in assets + watchlist:
            sym = item.get("symbol", "")
            atype = item.get("asset_type", "stock")
            if atype == "crypto":
                crypto_ids.add(sym.lower())
            else:
                stock_symbols.add(sym)

        results: dict = {"stocks": {}, "crypto": {}, "indices": {}, "meta": {}}

        def _fetch_stocks():
            """Fetch BR stock data from brapi.dev."""
            if not stock_symbols:
                return

            def _save_stock_quote(
                sym: str,
                name: str,
                price: float | None,
                previous_close: float | None,
                change: float | None,
                change_pct: float | None,
                volume: float | None,
                market_cap: float | None,
                high: float | None,
                low: float | None,
                updated: float | None = None,
                source: str = "unknown",
                is_stale: bool = False,
                captured_at: str | None = None,
                latency_ms: float | None = None,
            ) -> None:
                results["stocks"][sym] = {
                    "symbol": sym,
                    "name": name,
                    "price": price,
                    "previous_close": previous_close,
                    "change": change,
                    "change_pct": change_pct,
                    "volume": volume,
                    "market_cap": market_cap,
                    "high": high,
                    "low": low,
                    "updated": updated,
                    "source": source,
                    "is_stale": is_stale,
                    "captured_at": captured_at or now_iso,
                    "latency_ms": latency_ms,
                }
                _aid = repo.upsert_fin_asset({
                    "symbol": sym,
                    "name": name,
                    "asset_type": "stock",
                    "current_price": price,
                    "previous_close": previous_close,
                    "day_change": change,
                    "day_change_pct": change_pct,
                    "market_cap": market_cap,
                    "volume": volume,
                    "extra": {
                        "high": high,
                        "low": low,
                    },
                })
                if _aid and price:
                    repo.record_fin_asset_price(_aid, price, volume)

            brapi_token = str(app.config.get("BRAPI_TOKEN", "")).strip()
            if not brapi_token:
                brapi_token = repo.get_setting("brapi_token", "").strip()
                if brapi_token:
                    app.config["BRAPI_TOKEN"] = brapi_token

            for sym in sorted(stock_symbols):
                if not sym or sym.startswith("^"):
                    continue

                loaded = False

                # 1) Yahoo first to preserve brapi monthly quota.
                try:
                    y_started = time.perf_counter()
                    yresp = _http_get_with_retry(
                        "https://query1.finance.yahoo.com/v7/finance/quote",
                        params={"symbols": f"{sym}.SA"},
                        timeout=10,
                    )
                    y_latency_ms = round((time.perf_counter() - y_started) * 1000, 1)
                    _track_api_provider_latency("yahoo", "market-data", y_latency_ms)
                    _track_api_provider_usage("yahoo", bool(yresp.ok), "market-data")
                    if yresp.ok:
                        ydata = yresp.json()
                        items = ydata.get("quoteResponse", {}).get("result", [])
                        if items:
                            item = items[0]
                            price = item.get("regularMarketPrice")
                            if price is not None:
                                _save_stock_quote(
                                    sym=sym,
                                    name=item.get("longName")
                                    or item.get("shortName")
                                    or sym,
                                    price=price,
                                    previous_close=item.get(
                                        "regularMarketPreviousClose",
                                    ),
                                    change=item.get("regularMarketChange"),
                                    change_pct=item.get("regularMarketChangePercent"),
                                    volume=item.get("regularMarketVolume"),
                                    market_cap=item.get("marketCap"),
                                    high=item.get("regularMarketDayHigh"),
                                    low=item.get("regularMarketDayLow"),
                                    updated=item.get("regularMarketTime"),
                                    source="yahoo",
                                    is_stale=False,
                                    captured_at=now_iso,
                                    latency_ms=y_latency_ms,
                                )
                                loaded = True
                except Exception as exc:
                    _track_api_provider_usage("yahoo", False, "market-data")
                    logger.warning("Yahoo Finance fetch failed for %s: %s", sym, exc)

                if loaded:
                    continue

                # 2) brapi fallback only if quota/token available.
                if not brapi_token:
                    continue
                if not _try_reserve_brapi_call():
                    continue
                try:
                    params = {"fundamental": "true", "token": brapi_token}
                    b_started = time.perf_counter()
                    resp = _http_get_with_retry(
                        f"https://brapi.dev/api/quote/{sym}",
                        params=params,
                        timeout=10,
                    )
                    b_latency_ms = round((time.perf_counter() - b_started) * 1000, 1)
                    _track_api_provider_latency("brapi", "market-data", b_latency_ms)
                    _track_api_provider_usage("brapi", bool(resp.ok), "market-data")
                    if resp.ok:
                        data = resp.json()
                        items = data.get("results", [])
                        if not items:
                            continue
                        item = items[0]
                        price = item.get("regularMarketPrice")
                        if price is None:
                            continue
                        _save_stock_quote(
                            sym=sym,
                            name=item.get("longName", sym),
                            price=price,
                            previous_close=item.get("regularMarketPreviousClose"),
                            change=item.get("regularMarketChange"),
                            change_pct=item.get("regularMarketChangePercent"),
                            volume=item.get("regularMarketVolume"),
                            market_cap=item.get("marketCap"),
                            high=item.get("regularMarketDayHigh"),
                            low=item.get("regularMarketDayLow"),
                            updated=item.get("regularMarketTime"),
                            source="brapi",
                            is_stale=False,
                            captured_at=now_iso,
                            latency_ms=b_latency_ms,
                        )
                        _track_api_provider_fallback("brapi", "market-data")
                        loaded = True
                except Exception as exc:
                    _track_api_provider_usage("brapi", False, "market-data")
                    logger.warning("brapi.dev fetch failed for %s: %s", sym, exc)

                if loaded:
                    continue

                # 3) DB fallback — use last known price saved in fin_assets.
                db_asset = repo.get_fin_asset_by_symbol(sym)
                if db_asset and db_asset.get("current_price") is not None:
                    stale_price = float(db_asset["current_price"])
                    _track_api_provider_usage("db-fallback", True, "market-data")
                    _track_api_provider_fallback("db-fallback", "market-data")
                    results["stocks"][sym] = {
                        "symbol": sym,
                        "name": db_asset.get("name") or sym,
                        "price": stale_price,
                        "previous_close": db_asset.get("previous_close"),
                        "change": db_asset.get("day_change"),
                        "change_pct": db_asset.get("day_change_pct"),
                        "volume": db_asset.get("volume"),
                        "market_cap": db_asset.get("market_cap"),
                        "high": None,
                        "low": None,
                        "updated": None,
                        "stale": True,
                        "source": "db-fallback",
                        "is_stale": True,
                        "captured_at": db_asset.get("updated_at") or now_iso,
                        "latency_ms": None,
                    }
                    logger.info(
                        "DB fallback price used for %s: %.4f (stale)",
                        sym, stale_price,
                    )

        def _fetch_crypto():
            """Fetch crypto from CoinGecko."""
            if not crypto_ids:
                return
            try:
                ids_param = ",".join(crypto_ids)
                c_started = time.perf_counter()
                resp = _http_get_with_retry(
                    "https://api.coingecko.com/api/v3/simple/price",
                    params={
                        "ids": ids_param,
                        "vs_currencies": "brl,usd",
                        "include_24hr_change": "true",
                        "include_market_cap": "true",
                        "include_24hr_vol": "true",
                    },
                    timeout=10,
                )
                c_latency_ms = round((time.perf_counter() - c_started) * 1000, 1)
                _track_api_provider_latency(
                    "coingecko",
                    "market-data",
                    c_latency_ms,
                )
                _track_api_provider_usage("coingecko", bool(resp.ok), "market-data")
                if resp.ok:
                    data = resp.json()
                    for cid, info in data.items():
                        results["crypto"][cid] = {
                            "id": cid,
                            "price_brl": info.get("brl"),
                            "price_usd": info.get("usd"),
                            "change_24h": info.get("brl_24h_change"),
                            "market_cap": info.get("brl_market_cap"),
                            "volume_24h": info.get("brl_24h_vol"),
                            "source": "coingecko",
                            "is_stale": False,
                            "captured_at": now_iso,
                            "latency_ms": c_latency_ms,
                        }
                        _aid = repo.upsert_fin_asset({
                            "symbol": cid.upper(),
                            "name": cid.title(),
                            "asset_type": "crypto",
                            "currency": "BRL",
                            "current_price": info.get("brl"),
                            "day_change_pct": info.get("brl_24h_change"),
                            "market_cap": info.get("brl_market_cap"),
                            "volume": info.get("brl_24h_vol"),
                        })
                        _cprice = info.get("brl")
                        if _aid and _cprice:
                            repo.record_fin_asset_price(
                                _aid, _cprice,
                                info.get("brl_24h_vol"),
                            )
            except Exception as exc:
                _track_api_provider_usage("coingecko", False, "market-data")
                logger.warning("CoinGecko fetch failed: %s", exc)

        def _fetch_indices():
            """Fetch market indices via Yahoo first and brapi fallback for IBOV."""
            _INDEX_SYMBOLS = {
                "^BVSP": "Ibovespa",
                "^GSPC": "S&P 500",
                "^IXIC": "Nasdaq",
                "USDBRL=X": "Dólar (BRL)",
            }

            # Yahoo Finance v8 chart fallback — fetch all in parallel
            def _yahoo_index(yahoo_sym: str, default_name: str) -> None:
                try:
                    y_started = time.perf_counter()
                    yresp = _http_get_with_retry(
                        f"https://query2.finance.yahoo.com/v8/finance/chart/{yahoo_sym}",
                        params={"interval": "1d", "range": "1d"},
                        headers={"User-Agent": "Mozilla/5.0"},
                        timeout=8,
                    )
                    y_latency_ms = round((time.perf_counter() - y_started) * 1000, 1)
                    _track_api_provider_latency(
                        "yahoo",
                        "market-data",
                        y_latency_ms,
                    )
                    _track_api_provider_usage("yahoo", bool(yresp.ok), "market-data")
                    if not yresp.ok:
                        return
                    chart = yresp.json().get("chart", {}).get("result", [])
                    if not chart:
                        return
                    meta = chart[0].get("meta", {})
                    price = meta.get("regularMarketPrice")
                    prev_close = meta.get("chartPreviousClose") or meta.get(
                        "previousClose",
                    )
                    change = (
                        round(price - prev_close, 4)
                        if price is not None and prev_close
                        else None
                    )
                    change_pct = (
                        round(change / prev_close * 100, 4)
                        if change is not None and prev_close
                        else None
                    )
                    results["indices"][yahoo_sym] = {
                        "name": meta.get("longName")
                        or meta.get("shortName")
                        or default_name,
                        "price": price,
                        "change": change,
                        "change_pct": change_pct,
                        "source": "yahoo",
                        "is_stale": False,
                        "captured_at": now_iso,
                        "latency_ms": y_latency_ms,
                    }
                except Exception as exc:
                    _track_api_provider_usage("yahoo", False, "market-data")
                    logger.warning(
                        "Yahoo indices fetch failed for %s: %s", yahoo_sym, exc,
                    )

            with ThreadPoolExecutor(max_workers=4) as idx_executor:
                idx_futures = [
                    idx_executor.submit(_yahoo_index, sym, name)
                    for sym, name in _INDEX_SYMBOLS.items()
                ]
                for f in as_completed(idx_futures):
                    f.result()

            # brapi fallback only for IBOV when Yahoo is unavailable.
            if "^BVSP" in results["indices"]:
                return
            brapi_token = str(app.config.get("BRAPI_TOKEN", "")).strip()
            if not brapi_token:
                brapi_token = repo.get_setting("brapi_token", "").strip()
            if not brapi_token:
                return
            if not _try_reserve_brapi_call():
                return
            try:
                b_started = time.perf_counter()
                resp = _http_get_with_retry(
                    "https://brapi.dev/api/quote/%5EBVSP",
                    params={"token": brapi_token},
                    timeout=8,
                )
                b_latency_ms = round((time.perf_counter() - b_started) * 1000, 1)
                _track_api_provider_latency(
                    "brapi",
                    "market-data",
                    b_latency_ms,
                )
                _track_api_provider_usage("brapi", bool(resp.ok), "market-data")
                if resp.ok:
                    data = resp.json()
                    items = data.get("results", [])
                    if items:
                        item = items[0]
                        results["indices"]["^BVSP"] = {
                            "name": item.get("longName", "Ibovespa"),
                            "price": item.get("regularMarketPrice"),
                            "change": item.get("regularMarketChange"),
                            "change_pct": item.get("regularMarketChangePercent"),
                            "source": "brapi",
                            "is_stale": False,
                            "captured_at": now_iso,
                            "latency_ms": b_latency_ms,
                        }
                        _track_api_provider_fallback("brapi", "market-data")
            except Exception as exc:
                _track_api_provider_usage("brapi", False, "market-data")
                logger.warning("brapi indices fetch failed: %s", exc)

        # Run all fetches in parallel
        with ThreadPoolExecutor(max_workers=3) as executor:
            futures = [
                executor.submit(_fetch_stocks),
                executor.submit(_fetch_crypto),
                executor.submit(_fetch_indices),
            ]
            for f in as_completed(futures):
                try:
                    f.result()
                except Exception as exc:
                    logger.warning("market-data worker failed: %s", exc)

        if brapi_usage_delta > 0:
            final_usage = brapi_usage_start + brapi_usage_delta
            repo.set_setting(brapi_usage_key, str(final_usage))
        else:
            final_usage = brapi_usage_start

        remaining = max(0, brapi_limit - final_usage)
        results["meta"]["brapi"] = {
            "month": brapi_month_key,
            "usage": final_usage,
            "limit": brapi_limit,
            "reserve_pct": brapi_reserve_pct,
            "reserve_calls": brapi_reserve_calls,
            "usable_limit": brapi_usable_limit,
            "remaining_usable": max(0, brapi_usable_limit - final_usage),
            "per_request_cap": brapi_max_calls_per_request,
            "request_brapi_calls": brapi_usage_delta,
            "remaining": remaining,
            "degraded": final_usage >= brapi_usable_limit,
        }

        stale_stocks = sum(1 for q in results["stocks"].values() if q.get("is_stale"))
        stale_crypto = sum(1 for q in results["crypto"].values() if q.get("is_stale"))
        stale_indices = sum(1 for q in results["indices"].values() if q.get("is_stale"))
        stale_items = stale_stocks + stale_crypto + stale_indices
        results["meta"]["quality"] = {
            "captured_at": now_iso,
            "has_stale_data": stale_items > 0,
            "stale_items": stale_items,
            "stale_stocks": stale_stocks,
            "stale_crypto": stale_crypto,
            "stale_indices": stale_indices,
        }

        cache.set("finance:market", results, FINANCE_CACHE_TTLS["market"])
        return jsonify(results)

    # ── AI Financial Analysis ───────────────────────────────

    @app.post("/api/finance/ai-analysis")
    @limiter.limit("6/minute")
    @require_finance_key
    def finance_ai_analysis():
        """AI-powered financial analysis."""
        if not app.config.get("AI_LOCAL_ENABLED"):
            return jsonify({"error": "IA local não habilitada"}), 503

        body = request.get_json(silent=True) or {}
        user_msg = sanitize_text(
            str(body.get("message", "Analise meu portfólio")), 500,
        )
        analysis_type = str(body.get("type", "general"))

        # Build financial context
        summary = repo.get_fin_summary()
        currency = repo.list_currency_rates()
        goals = repo.list_fin_goals()
        transactions = repo.list_fin_transactions(limit=20)

        ctx_parts = []

        # Portfolio summary
        ctx_parts.append(
            f"Portfólio: valor total R${summary['current_value']:,.2f}, "
            f"investido R${summary['total_invested']:,.2f}, "
            f"P&L R${summary['total_pnl']:,.2f} ({summary['total_pnl_pct']:.1f}%)"
        )

        # Holdings
        for p in summary.get("portfolio", [])[:15]:
            current = (p.get("current_price") or 0) * p.get("quantity", 0)
            pnl = current - p.get("total_invested", 0)
            ctx_parts.append(
                f"  - {p['symbol']}: {p['quantity']:.2f} un × "
                f"R${p.get('current_price', 0):.2f} = R${current:,.2f} "
                f"(P&L R${pnl:,.2f})"
            )

        # Allocation
        alloc = summary.get("allocation", {})
        if alloc:
            parts = [f"{k}: R${v:,.2f}" for k, v in alloc.items()]
            ctx_parts.append(f"Alocação: {', '.join(parts)}")

        # Currency
        for c in currency:
            ctx_parts.append(
                f"Câmbio {c['pair']}: R${c['rate']:.4f} ({c.get('variation', 0):+.2f}%)"
            )

        # Goals
        for g in goals[:5]:
            pct = (
                (g["current_amount"] / g["target_amount"] * 100)
                if g["target_amount"] else 0
            )
            ctx_parts.append(
                f"Meta '{g['name']}': R${g['current_amount']:,.2f} / "
                f"R${g['target_amount']:,.2f} ({pct:.0f}%)"
            )

        # Recent transactions
        if transactions:
            tx_summary = []
            for t in transactions[:10]:
                tx_summary.append(
                    f"{t['tx_type'].upper()} {t['symbol']} "
                    f"{t['quantity']:.2f}×R${t['price']:.2f}"
                )
            ctx_parts.append(
                f"Últimas transações: {'; '.join(tx_summary)}"
            )

        system_prompt = (
            "Você é um consultor financeiro pessoal especializado no mercado brasileiro. "
            "Analise os dados financeiros do usuário e responda em português de forma clara e objetiva. "
            "Inclua análises de risco, diversificação, tendências e recomendações quando relevante. "
            "Use formatação markdown para organizar a resposta.\n\n"
            "Dados financeiros atuais:\n" + "\n".join(ctx_parts)
        )

        if analysis_type == "risk":
            system_prompt += (
                "\n\nFoco: análise detalhada de riscos, "
                "concentração, exposição setorial e volatilidade."
            )
        elif analysis_type == "allocation":
            system_prompt += (
                "\n\nFoco: análise de diversificação e alocação. "
                "Sugira rebalanceamento se necessário."
            )
        elif analysis_type == "performance":
            system_prompt += (
                "\n\nFoco: análise de performance, comparar com "
                "benchmarks como IBOV, CDI, IPCA."
            )

        try:
            ai_url = app.config["AI_LOCAL_URL"].rstrip("/")
            endpoint = app.config.get(
                "AI_LOCAL_LLAMA_CPP_CHAT_ENDPOINT",
                "/v1/chat/completions",
            )
            timeout_s = max(
                90,
                int(app.config.get("AI_LOCAL_TIMEOUT_SECONDS", 30)),
            )
            resp = http_requests.post(
                f"{ai_url}{endpoint}",
                json={
                    "model": app.config.get(
                        "AI_LOCAL_MODEL", "qwen2.5:7b-instruct",
                    ),
                    "messages": [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_msg},
                    ],
                    "max_tokens": 600,
                    "temperature": 0.7,
                },
                timeout=timeout_s,
            )
            if resp.ok:
                data = resp.json()
                choices = data.get("choices", [])
                answer = (
                    choices[0]["message"]["content"]
                    if choices else "Sem resposta"
                )
                return jsonify({"answer": answer, "type": analysis_type})
            return jsonify({
                "error": f"AI retornou {resp.status_code}",
            }), 502
        except Exception as exc:
            logger.warning("Finance AI analysis failed: %s", exc)
            return jsonify({"error": str(exc)}), 502

    # ── AI Chat Financeiro (com auto-cadastro) ────────────

    def _execute_ai_actions(
        actions: list[dict],
    ) -> list[dict]:
        """Execute structured actions returned by AI."""
        results: list[dict] = []
        for act in actions:
            act_type = act.get("action", "")
            try:
                if act_type == "add_asset":
                    symbol = sanitize_text(
                        str(act.get("symbol", "")).upper().strip(), 20,
                    )
                    if not symbol:
                        continue
                    aid = repo.upsert_fin_asset({
                        "symbol": symbol,
                        "name": sanitize_text(
                            str(act.get("name", symbol)), 100,
                        ),
                        "asset_type": _parse_asset_type(
                            str(act.get("asset_type", "stock")),
                        ),
                        "currency": sanitize_text(
                            str(act.get("currency", "BRL")).upper(), 10,
                        ),
                    })
                    results.append({
                        "action": "add_asset",
                        "symbol": symbol,
                        "id": aid,
                        "ok": True,
                    })

                elif act_type in ("add_transaction", "buy", "sell"):
                    symbol = sanitize_text(
                        str(act.get("symbol", "")).upper().strip(), 20,
                    )
                    if not symbol:
                        continue
                    # Ensure asset exists
                    asset = repo.get_fin_asset_by_symbol(symbol)
                    if not asset:
                        aid = repo.upsert_fin_asset({
                            "symbol": symbol,
                            "name": sanitize_text(
                                str(act.get("name", symbol)), 100,
                            ),
                            "asset_type": _parse_asset_type(
                                str(act.get("asset_type", "stock")),
                            ),
                            "currency": "BRL",
                        })
                    else:
                        aid = asset["id"]

                    qty = float(act.get("quantity", 0))
                    price = float(act.get("price", 0))
                    fees = float(act.get("fees", 0))
                    if qty <= 0 or price <= 0:
                        results.append({
                            "action": act_type,
                            "symbol": symbol,
                            "ok": False,
                            "error": "quantidade e preço obrigatórios",
                        })
                        continue

                    tx_type = "sell" if act_type == "sell" else (
                        _parse_tx_type(str(act.get("tx_type", "buy")))
                    )
                    tx_date = str(act.get("date", "")) or (
                        datetime.now().strftime("%Y-%m-%d")
                    )
                    total = qty * price + fees

                    tx_id = repo.add_fin_transaction({
                        "asset_id": aid,
                        "tx_type": tx_type,
                        "quantity": qty,
                        "price": price,
                        "total": total,
                        "fees": fees,
                        "notes": sanitize_text(
                            str(act.get("notes", "Via IA")), 500,
                        ),
                        "tx_date": tx_date,
                    })
                    _recalc_portfolio(repo, aid)
                    results.append({
                        "action": "add_transaction",
                        "symbol": symbol,
                        "tx_type": tx_type,
                        "quantity": qty,
                        "price": price,
                        "tx_id": tx_id,
                        "ok": True,
                    })

                elif act_type == "add_watchlist":
                    symbol = sanitize_text(
                        str(act.get("symbol", "")).upper().strip(), 20,
                    )
                    if not symbol:
                        continue
                    wid = repo.add_fin_watchlist({
                        "symbol": symbol,
                        "name": sanitize_text(
                            str(act.get("name", symbol)), 100,
                        ),
                        "asset_type": _parse_asset_type(
                            str(act.get("asset_type", "stock")),
                        ),
                        "target_price": (
                            float(act["target_price"])
                            if act.get("target_price") else None
                        ),
                        "alert_above": bool(act.get("alert_above")),
                        "notes": sanitize_text(
                            str(act.get("notes", "")), 500,
                        ),
                    })
                    results.append({
                        "action": "add_watchlist",
                        "symbol": symbol,
                        "id": wid,
                        "ok": True,
                    })

                elif act_type == "add_goal":
                    name = sanitize_text(
                        str(act.get("name", "")), 100,
                    )
                    target = float(act.get("target_amount", 0))
                    if not name or target <= 0:
                        continue
                    gid = repo.add_fin_goal({
                        "name": name,
                        "target_amount": target,
                        "current_amount": float(
                            act.get("current_amount", 0),
                        ),
                        "deadline": str(act.get("deadline", "")) or None,
                        "category": sanitize_text(
                            str(act.get("category", "savings")), 30,
                        ),
                        "notes": sanitize_text(
                            str(act.get("notes", "")), 500,
                        ),
                    })
                    results.append({
                        "action": "add_goal",
                        "name": name,
                        "id": gid,
                        "ok": True,
                    })

            except Exception as exc:
                results.append({
                    "action": act_type,
                    "ok": False,
                    "error": str(exc),
                })

        if results:
            _invalidate_financial_state_cache(include_market=True, include_dividends=True)
        return results

    @app.post("/api/finance/ai-chat")
    @limiter.limit("8/minute")
    @require_finance_key
    def finance_ai_chat():
        """Open-ended financial chat with AI + auto-register."""
        if not app.config.get("AI_LOCAL_ENABLED"):
            return jsonify({"error": "IA local não habilitada"}), 503

        body = request.get_json(silent=True) or {}
        user_msg = sanitize_text(str(body.get("message", "")), 800)
        if not user_msg:
            return jsonify({"error": "message obrigatório"}), 400

        # Build context
        summary = repo.get_fin_summary()
        currency = repo.list_currency_rates()
        goals = repo.list_fin_goals()
        assets = repo.list_fin_assets()

        ctx_parts = [
            f"Portfólio: R${summary['current_value']:,.2f} "
            f"(investido R${summary['total_invested']:,.2f}, "
            f"P&L {summary['total_pnl_pct']:+.1f}%).",
        ]
        for p in summary.get("portfolio", [])[:10]:
            ctx_parts.append(
                f"  {p['symbol']}: {p['quantity']:.2f}un "
                f"× R${p.get('current_price', 0):.2f}",
            )
        for c in currency:
            ctx_parts.append(
                f"  Câmbio {c['pair']}: R${c['rate']:.2f}",
            )
        for g in goals[:5]:
            pct = (
                (g["current_amount"] / g["target_amount"] * 100)
                if g["target_amount"] else 0
            )
            ctx_parts.append(
                f"  Meta '{g['name']}': {pct:.0f}% "
                f"(R${g['current_amount']:,.2f}/R${g['target_amount']:,.2f})",
            )
        asset_symbols = [a["symbol"] for a in assets[:30]]
        if asset_symbols:
            ctx_parts.append(
                f"  Ativos cadastrados: {', '.join(asset_symbols)}",
            )

        system_prompt = (
            "Você é um assistente financeiro pessoal brasileiro inteligente. "
            "Responda sempre em português de forma clara e objetiva.\n\n"
            "IMPORTANTE: Quando o usuário pedir para REGISTRAR, CADASTRAR, COMPRAR, "
            "VENDER, ADICIONAR um ativo, transação, item na watchlist ou meta financeira, "
            "você DEVE incluir um bloco JSON de ações no início da resposta no formato:\n"
            "```actions\n"
            '[{"action": "TIPO", ...campos}]\n'
            "```\n\n"
            "Tipos de ação disponíveis:\n"
            '- {"action":"add_asset","symbol":"PETR4","name":"Petrobras PN",'
            '"asset_type":"stock","currency":"BRL"}\n'
            '- {"action":"buy","symbol":"PETR4","quantity":100,"price":35.50,'
            '"date":"2026-01-15","notes":"..."}\n'
            '- {"action":"sell","symbol":"PETR4","quantity":50,"price":40.00,'
            '"date":"2026-01-15"}\n'
            '- {"action":"add_watchlist","symbol":"VALE3","name":"Vale SA",'
            '"asset_type":"stock","target_price":68.00,"alert_above":false}\n'
            '- {"action":"add_goal","name":"Reserva de emergência",'
            '"target_amount":50000,"current_amount":10000,'
            '"deadline":"2027-01-01","category":"savings"}\n\n'
            "Regras:\n"
            "- Múltiplas ações podem ser enviadas no mesmo bloco JSON (array).\n"
            "- Sempre confirme ao usuário o que foi feito após o bloco de ações.\n"
            "- asset_type pode ser: stock, fii, etf, crypto, fund, renda-fixa\n"
            "- Se o usuário não informar o preço ou quantidade, PERGUNTE antes de criar.\n"
            "- fees (taxas/corretagem) deve ser 0 a menos que o usuário informe explicitamente.\n"
            "- Se for apenas uma pergunta/análise, NÃO inclua o bloco actions.\n"
            "- Use markdown na resposta textual.\n\n"
            "Dados financeiros atuais do usuário:\n"
            + "\n".join(ctx_parts)
        )

        try:
            ai_url = app.config["AI_LOCAL_URL"].rstrip("/")
            endpoint = app.config.get(
                "AI_LOCAL_LLAMA_CPP_CHAT_ENDPOINT",
                "/v1/chat/completions",
            )
            resp = http_requests.post(
                f"{ai_url}{endpoint}",
                json={
                    "model": app.config.get(
                        "AI_LOCAL_MODEL", "qwen2.5:7b-instruct",
                    ),
                    "messages": [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_msg},
                    ],
                    "max_tokens": 700,
                    "temperature": 0.4,
                },
                timeout=app.config.get("AI_LOCAL_TIMEOUT_SECONDS", 45),
            )
            if not resp.ok:
                return jsonify({
                    "error": f"AI retornou {resp.status_code}",
                }), 502

            data = resp.json()
            choices = data.get("choices", [])
            raw_answer = (
                choices[0]["message"]["content"]
                if choices else "Sem resposta"
            )

            # Extract action blocks from AI response
            actions_executed: list[dict] = []
            clean_answer = raw_answer

            # Pattern: ```actions\n[...]\n```
            action_pattern = re.compile(
                r"```actions?\s*\n(.*?)\n```",
                re.DOTALL | re.IGNORECASE,
            )
            matches = action_pattern.findall(raw_answer)

            if not matches:
                # Also try inline JSON arrays [{"action":...}]
                json_pattern = re.compile(
                    r'\[\s*\{["\']action["\'].*?\}\s*\]',
                    re.DOTALL,
                )
                matches = json_pattern.findall(raw_answer)

            for match in matches:
                try:
                    parsed = json.loads(match)
                    if isinstance(parsed, dict):
                        parsed = [parsed]
                    if isinstance(parsed, list):
                        results = _execute_ai_actions(parsed)
                        actions_executed.extend(results)
                except (json.JSONDecodeError, TypeError):
                    pass

            # Remove action blocks from displayed answer
            clean_answer = action_pattern.sub("", clean_answer).strip()
            clean_answer = re.sub(
                r'\[\s*\{["\']action["\'].*?\}\s*\]',
                "",
                clean_answer,
                flags=re.DOTALL,
            ).strip()

            return jsonify({
                "answer": clean_answer,
                "actions": actions_executed,
            })

        except Exception as exc:
            return jsonify({"error": str(exc)}), 502

def _fmt_brl(value: float) -> str:
    """Format a float value as BRL currency string (server-side)."""
    try:
        v = float(value or 0)
        return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"

def _recalc_portfolio(repo: Repository, asset_id: int) -> None:
    """Recalculate portfolio position from all transactions."""
    txns = repo.list_fin_transactions(asset_id=asset_id, limit=9999)
    qty = 0.0
    total_cost = 0.0
    for t in reversed(txns):
        if t["tx_type"] == "buy":
            qty += t["quantity"]
            total_cost += t["total"]
        elif t["tx_type"] == "sell":
            if qty > 0:
                avg = total_cost / qty
                sell_qty = min(t["quantity"], qty)
                qty -= sell_qty
                total_cost -= avg * sell_qty

    if qty > 0:
        avg_price = total_cost / qty
        repo.upsert_fin_portfolio(asset_id, qty, avg_price, total_cost)
    else:
        repo.delete_fin_portfolio(asset_id)
